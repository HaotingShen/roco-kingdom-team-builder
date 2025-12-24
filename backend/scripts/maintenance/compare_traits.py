#!/usr/bin/env python3
"""
Compare traits from data_all.xlsx (Column G) with traits.json.
Identifies missing traits, description differences, and extra traits.
Generates text reports similar to compare_moves.py.
"""
import json
import sys
import zipfile
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

# File paths
EXCEL_PATH = Path("backend/data/data_all.xlsx")
JSON_PATH = Path("backend/data/traits.json")
REPORTS_DIR = Path("backend/data/trait_reports")


def load_json(path: Path) -> List[dict]:
    """Load JSON file."""
    try:
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"❌ Error: JSON file not found: {path}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"❌ Error: Invalid JSON format in {path}")
        print(f"   {e}")
        sys.exit(1)


def get_chinese_name(trait: dict) -> str:
    """Extract Chinese name from trait data."""
    return trait.get('localized', {}).get('zh', {}).get('name', '')


def get_chinese_description(trait: dict) -> str:
    """Extract Chinese description from trait data."""
    return trait.get('localized', {}).get('zh', {}).get('description', '')


def normalize_punctuation(text: str) -> str:
    """
    Convert English punctuation to Chinese punctuation in Chinese text.
    Also removes extra spaces after Chinese punctuation marks.

    English → Chinese:
    , → ，(comma)
    . → 。(period)
    : → ：(colon)
    ; → ；(semicolon)
    ! → ！(exclamation)
    ? → ？(question mark)
    ( → （(left paren)
    ) → ）(right paren)
    [ → 【(left bracket)
    ] → 】(right bracket)
    """
    if not text:
        return ''

    # Convert English punctuation to Chinese
    text = text.replace(',', '，')
    text = text.replace(';', '；')
    text = text.replace(':', '：')
    text = text.replace('!', '！')
    text = text.replace('?', '？')
    text = text.replace('(', '（')
    text = text.replace(')', '）')
    text = text.replace('[', '【')
    text = text.replace(']', '】')

    # Replace English periods with Chinese periods
    text = text.replace('. ', '。')
    text = text.replace('.', '。')

    # Remove spaces after Chinese punctuation marks
    text = text.replace('， ', '，')
    text = text.replace('。 ', '。')
    text = text.replace('： ', '：')
    text = text.replace('； ', '；')
    text = text.replace('！ ', '！')
    text = text.replace('？ ', '？')
    text = text.replace('） ', '）')
    text = text.replace('】 ', '】')

    # Ensure it ends with Chinese period
    if text and not text.endswith('。'):
        text = text + '。'

    return text


def find_trait_column(sheet) -> int:
    """
    Find column G by searching for header "✦ 特 性 ✦".

    Args:
        sheet: openpyxl worksheet object

    Returns:
        Column index (1-based, e.g., 7 for column G)

    Raises:
        ValueError: If trait column header not found
    """
    header_row = 1
    for col_idx in range(1, 21):  # Check first 20 columns
        cell_value = sheet.cell(row=header_row, column=col_idx).value
        if cell_value and "特 性" in str(cell_value):
            return col_idx
    raise ValueError("Trait column '✦ 特 性 ✦' not found in first 20 columns")


def parse_trait_cell(cell_value: str, row_idx: int = 0) -> Tuple[Optional[str], Optional[str]]:
    """
    Parse trait cell format: "{name}：{description}".

    Args:
        cell_value: Raw cell text (e.g., "氧循环：释放草系技能时，回复10%最大生命。")
        row_idx: Row index for warning messages

    Returns:
        (trait_name, trait_description) tuple, or (None, None) if invalid
    """
    if not cell_value or not isinstance(cell_value, str):
        return None, None

    # Split on Chinese colon
    if "：" not in cell_value:
        if row_idx and cell_value.strip():
            print(f"⚠️  Warning: Invalid trait format at row {row_idx}: {cell_value[:50]}")
        return None, None

    # Split on FIRST colon only (in case description contains colons)
    parts = cell_value.split("：", 1)
    if len(parts) != 2:
        return None, None

    name = parts[0].strip()
    description = parts[1].strip()

    if not name or not description:
        return None, None

    return name, description


def load_traits_from_excel(excel_path: Path) -> Dict[str, str]:
    """
    Load all traits from Excel file column G.

    Args:
        excel_path: Path to data_all.xlsx

    Returns:
        Dictionary mapping trait_name -> trait_description (both in Chinese)
    """
    try:
        workbook = openpyxl.load_workbook(excel_path, read_only=True)
    except FileNotFoundError:
        print(f"❌ Error: Excel file not found: {excel_path}")
        print(f"   Please ensure data_all.xlsx exists in backend/data/")
        sys.exit(1)
    except InvalidFileException:
        print(f"❌ Error: Invalid Excel file format: {excel_path}")
        print(f"   File may be corrupted or not a valid .xlsx file")
        sys.exit(1)
    except zipfile.BadZipFile:
        print(f"❌ Error: Excel file is corrupted: {excel_path}")
        sys.exit(1)

    try:
        sheet = workbook.active
        col_idx = find_trait_column(sheet)
    except ValueError as e:
        print(f"❌ Error: {e}")
        print(f"   Expected column header containing '✦ 特 性 ✦' or '特 性'")
        print(f"   Please verify Excel file structure")
        sys.exit(1)

    excel_traits = {}
    seen_names = set()
    duplicates = []
    skipped_count = 0

    # Iterate rows starting from row 2 (skip header)
    for row_idx in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_idx, column=col_idx).value

        if not cell_value:
            skipped_count += 1
            continue

        name, description = parse_trait_cell(cell_value, row_idx)

        if not name or not description:
            skipped_count += 1
            continue

        if name in seen_names:
            duplicates.append((name, row_idx))
            continue  # Skip duplicate, keep first occurrence

        seen_names.add(name)
        excel_traits[name] = description

    if duplicates:
        print(f"⚠️  Warning: {len(duplicates)} duplicate trait names in Excel:")
        for name, row in duplicates[:5]:
            print(f"   - {name} (row {row})")
        if len(duplicates) > 5:
            print(f"   ... and {len(duplicates) - 5} more")

    if skipped_count > 0:
        print(f"ℹ️  Skipped {skipped_count} empty or invalid cells")

    return excel_traits


def compare_traits(excel_traits: Dict[str, str],
                  json_traits: List[dict]) -> Tuple[List[dict], List[dict], List[dict]]:
    """
    Compare traits from Excel and JSON.

    Args:
        excel_traits: Dictionary from load_traits_from_excel()
        json_traits: List from load_json()

    Returns:
        (missing_traits, different_traits, extra_traits)
    """
    # Build lookup map from JSON: {chinese_name: trait_data}
    json_traits_map = {}
    seen_names = set()
    duplicates = []

    for trait in json_traits:
        zh_name = get_chinese_name(trait)
        if not zh_name:
            continue

        if zh_name in seen_names:
            duplicates.append(zh_name)

        seen_names.add(zh_name)
        json_traits_map[zh_name] = trait

    if duplicates:
        print(f"⚠️  Warning: {len(duplicates)} duplicate trait names in JSON:")
        for name in duplicates[:5]:
            print(f"   - {name}")
        if len(duplicates) > 5:
            print(f"   ... and {len(duplicates) - 5} more")

    # Find missing traits (in Excel, not in JSON)
    missing = []
    for excel_name, excel_desc in excel_traits.items():
        if excel_name not in json_traits_map:
            missing.append({
                'chinese_name': excel_name,
                'excel_description': normalize_punctuation(excel_desc)
            })

    # Find different descriptions (in both, but desc differs)
    different = []
    for zh_name, trait_data in json_traits_map.items():
        if zh_name in excel_traits:
            json_desc = get_chinese_description(trait_data)
            excel_desc = excel_traits[zh_name]

            # Normalize both
            json_norm = normalize_punctuation(json_desc)
            excel_norm = normalize_punctuation(excel_desc)

            # Compare normalized versions
            if json_norm != excel_norm:
                different.append({
                    'chinese_name': zh_name,
                    'english_name': trait_data.get('name', ''),
                    'json_description': json_desc,  # original
                    'excel_description': excel_desc,  # original
                    'json_normalized': json_norm,
                    'excel_normalized': excel_norm
                })

    # Find extra traits (in JSON, not in Excel)
    extra = []
    for zh_name, trait_data in json_traits_map.items():
        if zh_name not in excel_traits:
            extra.append({
                'chinese_name': zh_name,
                'english_name': trait_data.get('name', ''),
                'json_description': get_chinese_description(trait_data)
            })

    return missing, different, extra


def generate_reports():
    """Generate comparison reports."""
    # Create reports directory
    REPORTS_DIR.mkdir(exist_ok=True)

    print("🔍 Comparing traits from Excel with traits.json...")
    print()

    # Load data
    excel_traits = load_traits_from_excel(EXCEL_PATH)
    json_traits = load_json(JSON_PATH)

    if not excel_traits:
        print("⚠️  Warning: No traits found in Excel file")
        print("   Column may be empty or all rows have invalid format")

    if not json_traits:
        print("⚠️  Warning: traits.json is empty or has no valid entries")

    print()

    # Run comparison
    missing, different, extra = compare_traits(excel_traits, json_traits)

    # Report 1: Missing traits
    print(f"📋 Report 1: Missing Traits ({len(missing)} traits)")
    print("=" * 80)
    if missing:
        report_path = REPORTS_DIR / "missing_traits.txt"
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("MISSING TRAITS (in Excel but not in traits.json)\n")
            f.write("=" * 80 + "\n\n")
            for i, trait in enumerate(missing, 1):
                line = f"{i}. {trait['chinese_name']}\n"
                line += f"   Description (ZH): {trait['excel_description']}\n\n"
                f.write(line)
                print(f"  {i}. {trait['chinese_name']}")
        print(f"\n✅ Saved to: {report_path}")
    else:
        print("  ✅ No missing traits!")
    print()

    # Report 2: Different descriptions
    print(f"📋 Report 2: Different Descriptions ({len(different)} traits)")
    print("=" * 80)
    if different:
        report_path = REPORTS_DIR / "different_descriptions.txt"
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("DIFFERENT DESCRIPTIONS (traits exist in both, but descriptions differ)\n")
            f.write("=" * 80 + "\n\n")
            for i, trait in enumerate(different, 1):
                line = f"{i}. {trait['chinese_name']} ({trait['english_name']})\n"
                line += f"   Excel (normalized):  {trait['excel_normalized']}\n"
                line += f"   JSON (normalized):   {trait['json_normalized']}\n"
                line += "\n"
                line += f"   Excel (original):    {trait['excel_description']}\n"
                line += f"   JSON (original):     {trait['json_description']}\n\n"
                f.write(line)
                print(f"  {i}. {trait['chinese_name']} - description differs")
        print(f"\n✅ Saved to: {report_path}")
    else:
        print("  ✅ No different descriptions!")
    print()

    # Report 3: Extra traits
    print(f"📋 Report 3: Extra Traits ({len(extra)} traits)")
    print("=" * 80)
    if extra:
        report_path = REPORTS_DIR / "extra_traits.txt"
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("EXTRA TRAITS (in traits.json but not in Excel)\n")
            f.write("=" * 80 + "\n\n")
            for i, trait in enumerate(extra, 1):
                line = f"{i}. {trait['chinese_name']} ({trait['english_name']})\n"
                line += f"   Description (ZH): {trait['json_description']}\n\n"
                f.write(line)
                print(f"  {i}. {trait['chinese_name']} ({trait['english_name']})")
        print(f"\n✅ Saved to: {report_path}")
    else:
        print("  ✅ No extra traits!")
    print()

    # Summary
    print("=" * 80)
    print("📊 SUMMARY")
    print("=" * 80)
    print(f"Total traits in Excel: {len(excel_traits)}")
    print(f"Total traits in traits.json: {len(json_traits)}")
    print(f"Missing traits (in Excel, not in JSON): {len(missing)}")
    print(f"Different descriptions: {len(different)}")
    print(f"Extra traits (in JSON, not in Excel): {len(extra)}")
    print()
    print(f"📁 Reports saved to: {REPORTS_DIR}/")
    print()


if __name__ == "__main__":
    generate_reports()
