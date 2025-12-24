"""
Fix monster base stats mismatches between Excel source and monsters.json.

This script:
1. Reads base stats from Excel (source of truth)
2. Matches monsters with JSON entries
3. Updates stat fields where there are mismatches
4. Creates a backup before saving changes

Usage:
    python -m backend.scripts.maintenance.fix_monster_stats          # Interactive mode
    python -m backend.scripts.maintenance.fix_monster_stats --dry-run # Dry run (no changes)
"""

import json
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# ANSI color codes
class Colors:
    RESET = "\033[0m"
    BOLD = "\033[1m"
    RED = "\033[91m"
    GREEN = "\033[92m"
    YELLOW = "\033[93m"
    BLUE = "\033[94m"
    CYAN = "\033[96m"


class MonsterStatsFixer:
    def __init__(self, data_dir: Path):
        self.data_dir = data_dir
        self.excel_path = data_dir / "data_all.xlsx"
        self.json_path = data_dir / "monsters.json"

        self.excel_monsters = {}
        self.json_monsters = []
        self.changes = []

    def load_excel_stats(self):
        """Load monster base stats from Excel."""
        print(f"{Colors.BOLD}Loading Excel data...{Colors.RESET}")
        wb = load_workbook(self.excel_path, read_only=True, data_only=True)
        ws = wb['基础属性']

        # First pass: identify leader monsters with multiple forms
        # A leader has multiple forms if it appears multiple times consecutively
        multi_form_leaders = set()
        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]
            monster_name = row[1].value
            stage = row[2].value

            if not monster_name or stage != '首领形态':
                continue

            # Check if next row(s) have the same leader monster name
            for next_row_idx in range(row_idx + 1, min(row_idx + 10, ws.max_row + 1)):
                next_row = ws[next_row_idx]
                next_name = next_row[1].value
                next_stage = next_row[2].value

                if next_name == monster_name and next_stage == '首领形态':
                    # Found another instance of the same leader → multi-form leader
                    multi_form_leaders.add(monster_name)
                    break
                elif next_stage == '首领形态' and next_name != monster_name:
                    # Different leader encountered, stop searching
                    break

        # Track previous row to determine leader forms
        prev_monster_name = None

        # Second pass: process each row with multi-form leader logic
        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]

            monster_name = row[1].value  # Column 2
            if not monster_name:
                continue

            stage = row[2].value  # Column 3

            # For leader monsters with multiple forms, append form from previous evolution stage
            # Only apply this logic if:
            # 1. Current monster is a leader (stage == "首领形态")
            # 2. Leader has multiple forms (appears multiple times)
            # 3. Previous row had a monster name
            full_name = monster_name
            if (stage == "首领形态" and
                monster_name in multi_form_leaders and
                prev_monster_name):

                # Special case: 棋契陛下 uses the FULL previous monster name as form
                # (e.g., "棋齐垒-黑子" instead of just "黑子")
                if monster_name == "棋契陛下":
                    full_name = f"{monster_name}-{prev_monster_name}"
                elif "-" in prev_monster_name:
                    # Normal case: extract just the form suffix (everything after the first '-')
                    form_part = prev_monster_name.split("-", 1)[1]
                    full_name = f"{monster_name}-{form_part}"

            # Extract base stats (columns 8-13)
            base_hp = row[7].value       # Column 8: 生命
            base_phy_atk = row[8].value  # Column 9: 物攻
            base_mag_atk = row[9].value  # Column 10: 魔攻
            base_phy_def = row[10].value # Column 11: 物防
            base_mag_def = row[11].value # Column 12: 魔防
            base_spd = row[12].value     # Column 13: 速度

            # Parse full_name to get base + form for lookup key
            if "-" in full_name:
                parts = full_name.split("-", 1)
                zh_base = parts[0].strip()
                zh_form = parts[1].strip()
            else:
                zh_base = full_name
                zh_form = None

            # Store with composite key
            key = (zh_base, zh_form)
            self.excel_monsters[key] = {
                "base_hp": base_hp,
                "base_phy_atk": base_phy_atk,
                "base_mag_atk": base_mag_atk,
                "base_phy_def": base_phy_def,
                "base_mag_def": base_mag_def,
                "base_spd": base_spd,
                "excel_name": full_name,
                "row": row_idx
            }

            # Update previous monster name for next iteration
            if monster_name:
                prev_monster_name = monster_name

        print(f"  Loaded {len(self.excel_monsters)} monsters from Excel")
        print(f"  Detected {len(multi_form_leaders)} multi-form leaders\n")

    def load_json_monsters(self):
        """Load monsters from JSON."""
        print(f"{Colors.BOLD}Loading JSON data...{Colors.RESET}")
        with open(self.json_path, 'r', encoding='utf-8') as f:
            self.json_monsters = json.load(f)
        print(f"  Loaded {len(self.json_monsters)} monsters from JSON\n")

    def fix_stats(self):
        """Fix base stats mismatches."""
        print(f"{Colors.BOLD}Checking and fixing base stats mismatches...{Colors.RESET}\n")

        fixed_count = 0
        skipped_count = 0
        total_stat_changes = 0
        stat_fields = [
            "base_hp",
            "base_phy_atk",
            "base_mag_atk",
            "base_phy_def",
            "base_mag_def",
            "base_spd"
        ]

        for monster in self.json_monsters:
            # Get JSON monster identifiers
            zh_base = monster.get("localized", {}).get("zh", {}).get("name")
            zh_form = monster.get("localized", {}).get("zh", {}).get("form")
            english_name = monster.get("name")

            if not zh_base:
                skipped_count += 1
                continue

            # Look up Excel data
            key = (zh_base, zh_form)
            excel_data = self.excel_monsters.get(key)

            if not excel_data:
                # Monster exists in JSON but not in Excel - skip
                skipped_count += 1
                continue

            # Check for mismatches in any stat field
            has_mismatch = False
            field_changes = []

            for field in stat_fields:
                json_value = monster.get(field)
                excel_value = excel_data.get(field)

                # Handle None/null values
                if excel_value is None:
                    continue

                if json_value != excel_value:
                    has_mismatch = True
                    field_changes.append({
                        "field": field,
                        "old": json_value,
                        "new": excel_value
                    })

            if has_mismatch:
                # Print header for this monster
                print(f"{Colors.YELLOW}Fixing {english_name} ({zh_base}{'-'+zh_form if zh_form else ''}):{Colors.RESET}")

                # Record the change
                change = {
                    "english_name": english_name,
                    "zh_name": zh_base,
                    "zh_form": zh_form,
                    "excel_row": excel_data["row"],
                    "changes": []
                }

                # Apply and print each stat change
                for field_change in field_changes:
                    field = field_change["field"]
                    old_val = field_change["old"]
                    new_val = field_change["new"]

                    # Format field name for display
                    display_name = field.replace("base_", "").replace("_", " ").title()

                    print(f"  {display_name:12s}: {Colors.RED}{old_val:>3}{Colors.RESET} → {Colors.GREEN}{new_val:>3}{Colors.RESET}")

                    # Update the monster
                    monster[field] = new_val

                    # Record the change
                    change["changes"].append(field_change)
                    total_stat_changes += 1

                self.changes.append(change)
                fixed_count += 1
                print()

        print(f"{Colors.BOLD}Summary:{Colors.RESET}")
        print(f"  Fixed: {Colors.GREEN}{fixed_count}{Colors.RESET} monsters ({total_stat_changes} stat field{'' if total_stat_changes == 1 else 's'})")
        print(f"  Skipped: {Colors.BLUE}{skipped_count}{Colors.RESET} monsters (not in Excel)")
        print()

    def save_backup(self):
        """Create backup of original JSON."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self.data_dir / f"monsters.json.backup_{timestamp}"

        print(f"{Colors.BOLD}Creating backup...{Colors.RESET}")
        with open(self.json_path, 'r', encoding='utf-8') as f:
            content = f.read()
        with open(backup_path, 'w', encoding='utf-8') as f:
            f.write(content)
        print(f"  Backup saved: {Colors.CYAN}{backup_path.name}{Colors.RESET}\n")

    def save_json(self):
        """Save updated JSON."""
        print(f"{Colors.BOLD}Saving updated monsters.json...{Colors.RESET}")
        with open(self.json_path, 'w', encoding='utf-8') as f:
            json.dump(self.json_monsters, f, ensure_ascii=False, indent=2)
        print(f"  {Colors.GREEN}✓{Colors.RESET} Saved successfully\n")

    def save_change_log(self):
        """Save detailed change log."""
        if not self.changes:
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_path = self.data_dir / f"stat_fixes_log_{timestamp}.json"

        print(f"{Colors.BOLD}Saving change log...{Colors.RESET}")
        with open(log_path, 'w', encoding='utf-8') as f:
            json.dump({
                "timestamp": timestamp,
                "total_changes": len(self.changes),
                "changes": self.changes
            }, f, ensure_ascii=False, indent=2)
        print(f"  Change log saved: {Colors.CYAN}{log_path.name}{Colors.RESET}\n")

    def run(self, dry_run=False):
        """Run the fix process."""
        print(f"\n{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.RESET}")
        print(f"{Colors.BOLD}{Colors.CYAN}Monster Base Stats Fixer{Colors.RESET}")
        if dry_run:
            print(f"{Colors.BOLD}{Colors.YELLOW}(DRY RUN - No changes will be saved){Colors.RESET}")
        print(f"{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.RESET}\n")

        # Load data
        self.load_excel_stats()
        self.load_json_monsters()

        # Fix stats
        self.fix_stats()

        if not self.changes:
            print(f"{Colors.GREEN}✓ No stat mismatches found - monsters.json is already correct!{Colors.RESET}\n")
            return

        # Show summary
        print(f"{Colors.BOLD}{Colors.YELLOW}Found {len(self.changes)} monster(s) with stat mismatches.{Colors.RESET}")

        if dry_run:
            print(f"\n{Colors.YELLOW}Dry run complete - no changes saved{Colors.RESET}\n")
            return

        # Ask for confirmation
        response = input(f"{Colors.BOLD}Apply these fixes? (yes/no): {Colors.RESET}").strip().lower()

        if response not in ['yes', 'y']:
            print(f"\n{Colors.YELLOW}Aborted - no changes made{Colors.RESET}\n")
            return

        # Save backup and apply changes
        print()
        self.save_backup()
        self.save_json()
        self.save_change_log()

        print(f"{Colors.BOLD}{Colors.GREEN}{'='*80}{Colors.RESET}")
        print(f"{Colors.BOLD}{Colors.GREEN}✓ Base stats fixes applied successfully!{Colors.RESET}")
        print(f"{Colors.BOLD}{Colors.GREEN}{'='*80}{Colors.RESET}\n")


def main():
    # Parse arguments
    dry_run = "--dry-run" in sys.argv or "-n" in sys.argv

    # Determine data directory
    script_dir = Path(__file__).parent
    data_dir = script_dir.parent.parent / "data"

    if not data_dir.exists():
        print(f"{Colors.RED}Error: Data directory not found: {data_dir}{Colors.RESET}")
        return 1

    fixer = MonsterStatsFixer(data_dir)
    fixer.run(dry_run=dry_run)

    return 0


if __name__ == "__main__":
    exit(main())
