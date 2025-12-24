"""
Reorder monsters.json to match the species/family order from data_all.xlsx.

This script:
1. Reads the Excel file to determine species order (by first occurrence)
2. Groups JSON entries by species while preserving internal order
3. Reorders species groups to match Excel order
4. Appends species not found in Excel to the end
"""

import json
import shutil
from pathlib import Path
from collections import OrderedDict
import openpyxl


def main():
    # Paths
    script_dir = Path(__file__).parent.parent
    excel_path = script_dir / "data" / "data_all.xlsx"
    json_path = script_dir / "data" / "monsters.json"
    backup_path = script_dir / "data" / "monsters.json.backup"

    print("=" * 60)
    print("Reordering monsters.json by Species/Family Order")
    print("=" * 60)

    # ========== Phase 1: Build Species Order from Excel ==========
    print("\n[Phase 1] Reading Excel file...")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['基础属性']

    # Extract all monster names from Excel (column B, skip header)
    excel_names = []
    for row_num in range(2, ws.max_row + 1):
        name = ws.cell(row=row_num, column=2).value
        if name:
            excel_names.append(name)

    print(f"Found {len(excel_names)} monster names in Excel")

    # ========== Phase 2: Group JSON Entries by Species ==========
    print("\n[Phase 2] Reading monsters.json...")
    with open(json_path, 'r', encoding='utf-8') as f:
        monsters = json.load(f)

    print(f"Found {len(monsters)} total entries in JSON")

    # Create name-to-species mapping (handles both base names and forms)
    name_to_species = {}
    for m in monsters:
        zh = m.get('localized', {}).get('zh', {})
        zh_name = zh.get('name', '')
        zh_form = zh.get('form')
        species = m['species']

        # Map base name to species
        if zh_name:
            name_to_species[zh_name] = species

        # Map full name with form (for entries like "丢丢-草地附近的样子")
        if zh_form:
            full_name = f'{zh_name}-{zh_form}'
            name_to_species[full_name] = species

    # Group JSON entries by species (preserving order within each group)
    species_groups = OrderedDict()
    for m in monsters:
        species = m['species']
        if species not in species_groups:
            species_groups[species] = []
        species_groups[species].append(m)

    print(f"Grouped into {len(species_groups)} species")

    # Determine species order from Excel (first occurrence)
    print("\n[Phase 3] Building species order from Excel...")
    species_order = []
    seen_species = set()

    for excel_name in excel_names:
        species = name_to_species.get(excel_name)
        if species and species not in seen_species:
            species_order.append(species)
            seen_species.add(species)

    print(f"Species order determined: {len(species_order)} species from Excel")
    print(f"First 10 species: {' → '.join(species_order[:10])}")

    # ========== Phase 3: Reorder Groups ==========
    print("\n[Phase 4] Reordering species groups...")
    reordered_monsters = []
    processed_species = set()

    # Add species in Excel order
    for species in species_order:
        if species in species_groups:
            group = species_groups[species]
            reordered_monsters.extend(group)
            processed_species.add(species)

    excel_entries_count = len(reordered_monsters)
    print(f"✓ Processed {len(processed_species)} species from Excel ({excel_entries_count} entries)")

    # Add remaining species not in Excel
    remaining_species = [s for s in species_groups.keys() if s not in processed_species]
    remaining_entries_count = 0

    for species in remaining_species:
        group = species_groups[species]
        reordered_monsters.extend(group)
        remaining_entries_count += len(group)

    print(f"✓ Added {len(remaining_species)} remaining species to end ({remaining_entries_count} entries)")

    if remaining_species:
        print(f"  Species not in Excel: {', '.join(remaining_species[:10])}")
        if len(remaining_species) > 10:
            print(f"  ... and {len(remaining_species) - 10} more")

    # ========== Validation ==========
    print("\n[Validation]")
    assert len(reordered_monsters) == len(monsters), \
        f"Entry count mismatch! Original: {len(monsters)}, Reordered: {len(reordered_monsters)}"
    print(f"✓ All {len(monsters)} entries preserved")

    assert len(set(m['species'] for m in reordered_monsters)) == len(species_groups), \
        "Species group count mismatch!"
    print(f"✓ All {len(species_groups)} species groups preserved")

    # ========== Backup and Write ==========
    print("\n[Output]")
    print(f"Creating backup: {backup_path.name}")
    shutil.copy(json_path, backup_path)

    print(f"Writing reordered JSON to: {json_path.name}")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(reordered_monsters, f, ensure_ascii=False, indent=2)

    # ========== Summary ==========
    print("\n" + "=" * 60)
    print("✓ Reordering complete!")
    print("=" * 60)
    print(f"Total entries: {len(reordered_monsters)}")
    print(f"Total species: {len(species_groups)}")
    print(f"Species from Excel: {len(species_order)} (entries 1-{excel_entries_count})")
    print(f"Species not in Excel: {len(remaining_species)} (entries {excel_entries_count + 1}-{len(reordered_monsters)})")
    print()
    print("First 4 species in reordered file:")
    current_idx = 0
    for i, species in enumerate(species_order[:4], 1):
        group = species_groups[species]
        names = [m.get('localized', {}).get('zh', {}).get('name', 'N/A') for m in group]
        print(f"  {i}. {species}: {', '.join(names)} ({len(group)} entries)")
        current_idx += len(group)
    print()


if __name__ == "__main__":
    main()
