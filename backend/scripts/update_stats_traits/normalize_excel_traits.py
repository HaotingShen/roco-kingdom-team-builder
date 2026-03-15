"""
Normalize punctuation in the trait column (column 10) of data_stats_traits_formal.xlsx.

Rules applied to the DESCRIPTION part only (after the first "：" or ":"):
  - English punctuation → Chinese equivalents (, → ，  . → 。  : → ：  etc.)
  - Remove spaces after Chinese punctuation
  - Ensure description ends with 。
  - Colon separator is normalized to Chinese full-width ：

The trait NAME (before the colon) is left untouched.

Output is saved as data_stats_traits_formal_normalized.xlsx (cannot overwrite NTFS
files in-place from WSL; rename/replace manually if needed).

Usage (from project root):
    python -m backend.scripts.update_stats_traits.normalize_excel_traits
"""

import sys
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Punctuation normaliser
# ---------------------------------------------------------------------------

def normalize_punctuation(text: str) -> str:
    if not text:
        return ""
    text = text.replace(",", "，")
    text = text.replace(";", "；")
    text = text.replace("!", "！")
    text = text.replace("?", "？")
    text = text.replace("(", "（")
    text = text.replace(")", "）")
    text = text.replace("[", "【")
    text = text.replace("]", "】")
    text = text.replace(". ", "。")
    text = text.replace(".", "。")
    # Remove space after Chinese punctuation
    for p in "，。；！？）】":
        text = text.replace(p + " ", p)
    if text and not text.endswith("。"):
        text += "。"
    return text


def normalize_trait_cell(cell_value: str) -> str | None:
    """
    Parse "名称：描述" (or "名称:描述"), normalize the description,
    return the reconstructed cell string with a Chinese colon separator.
    Returns None if the cell cannot be parsed.
    """
    if not cell_value or not isinstance(cell_value, str):
        return None

    # Try Chinese colon first, then English colon
    for sep in ("：", ":"):
        if sep in cell_value:
            name, desc = cell_value.split(sep, 1)
            name = name.strip()
            desc = desc.strip()
            if name and desc:
                normalized_desc = normalize_punctuation(desc)
                return f"{name}：{normalized_desc}"

    return None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run():
    root = Path(__file__).parent.parent.parent.parent
    data_dir = root / "backend" / "data"
    excel_path = data_dir / "data_stats_traits_formal.xlsx"
    out_path   = data_dir / "data_stats_traits_formal_normalized.xlsx"

    print(f"Loading {excel_path.name} …")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    changed = 0
    skipped = 0

    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=10)
        raw = cell.value
        if not raw:
            continue

        normalized = normalize_trait_cell(str(raw))
        if normalized is None:
            print(f"  ⚠  Row {row_idx}: could not parse trait cell — skipped")
            skipped += 1
            continue

        if normalized != str(raw):
            cell.value = normalized
            changed += 1

    print(f"\n  {changed} cells normalized, {skipped} skipped")
    wb.save(out_path)
    print(f"Saved: {out_path.name}")
    print("(Rename to data_stats_traits_formal.xlsx to replace the original)")


if __name__ == "__main__":
    run()
