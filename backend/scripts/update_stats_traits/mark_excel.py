"""
Mark data_stats_traits_formal.xlsx with match status columns.

Reads each monster row from the Excel and cross-references against monsters.json,
then writes classification columns (11-14) back to the Excel WITHOUT modifying
any of the original 10 data columns.

Match types written to Column 11:
  EXACT_SINGLE       - Name matches exactly one zh_name group (all forms share the same stats row)
  EXACT_MULTI_FORM   - Name matches a zh_name that has multiple distinct forms; needs manual review
                       before updating, because the Excel doesn't specify which form(s) to target
  PARENTHETICAL      - Name contains a parenthetical hint e.g. "圣代甜甜 (抹茶)"; skipped from
                       auto-update because form matching is ambiguous
  NEW_MONSTER        - Name not found anywhere in monsters.json; needs to be added first

Column 12: Form count  - how many monsters.json entries matched by zh_name
Column 13: Form names  - semicolon-separated list of zh_form values (or "default" when None)
Column 14: Notes       - human-readable annotation

Usage (from project root):
    python -m backend.scripts.update_stats_traits.mark_excel
    python -m backend.scripts.update_stats_traits.mark_excel --dry-run  # print only, no write
"""

import json
import re
import sys
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Colour palette for status cells
# ---------------------------------------------------------------------------
FILL = {
    "EXACT_SINGLE":     PatternFill("solid", fgColor="C6EFCE"),   # green
    "EXACT_MULTI_FORM": PatternFill("solid", fgColor="FFEB9C"),   # yellow
    "PARENTHETICAL":    PatternFill("solid", fgColor="FFEB9C"),   # yellow
    "NEW_MONSTER":      PatternFill("solid", fgColor="FFC7CE"),   # red
}
FONT_BOLD = Font(bold=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_json(path: Path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def parse_excel_name(raw: str):
    """
    Split a raw Excel name into (zh_name, form_hint).

    Examples:
        "圣代甜甜 (抹茶)"  → ("圣代甜甜", "抹茶")
        "春花兔"           → ("春花兔", None)
    """
    raw = raw.strip()
    m = re.match(r'^(.+?)\s*[（(](.+?)[)）]\s*$', raw)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return raw, None


def build_monster_index(monsters: list) -> dict:
    """
    Returns: { zh_name: [ {en_name, zh_form, ...monster_entry}, ... ] }
    zh_form is None when the monster has no form (default).
    """
    index = defaultdict(list)
    for m in monsters:
        zh = m.get("localized", {}).get("zh", {})
        zh_name = zh.get("name")
        if not zh_name:
            continue
        index[zh_name].append({
            "en_name":  m.get("name"),
            "zh_form":  zh.get("form"),   # None when no form
            "trait":    m.get("trait"),
        })
    return index


def classify_row(zh_name: str, form_hint, index: dict):
    """
    Determine the match type for one Excel row.

    Returns (match_type, matched_entries, notes)
    """
    entries = index.get(zh_name)

    # ── Completely unknown name ─────────────────────────────────────────────
    if not entries:
        return "NEW_MONSTER", [], "Not found in monsters.json — add this monster first"

    # ── Parenthetical form hint ─────────────────────────────────────────────
    if form_hint is not None:
        # Try to find entries whose zh_form contains the hint
        matched = [e for e in entries if e["zh_form"] and form_hint in e["zh_form"]]
        if not matched:
            note = (
                f"Hint '{form_hint}' not found in any zh_form of '{zh_name}' "
                f"({', '.join(e['zh_form'] or 'default' for e in entries)}) — "
                "form names may have changed; review manually"
            )
        else:
            forms_str = "; ".join(e["zh_form"] for e in matched)
            note = f"Hint '{form_hint}' matched: {forms_str} — skipped from auto-update; review manually"
        return "PARENTHETICAL", matched, note

    # ── Exact name match ────────────────────────────────────────────────────
    # Check whether the monster actually has multiple distinct forms
    forms = [e["zh_form"] for e in entries]
    has_real_forms = any(f is not None for f in forms)

    if has_real_forms:
        note = (
            f"Monster has {len(entries)} form(s) in monsters.json — "
            "Excel provides one stat row for all forms; confirm stats are shared "
            "before running update_stats"
        )
        return "EXACT_MULTI_FORM", entries, note
    else:
        # Single or default-form monster
        en = entries[0]["en_name"]
        note = f"Matched → {en}"
        return "EXACT_SINGLE", entries, note


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(dry_run: bool = False):
    root = Path(__file__).parent.parent.parent.parent   # project root
    data_dir = root / "backend" / "data"
    excel_path = data_dir / "data_stats_traits_formal.xlsx"
    monsters_path = data_dir / "monsters.json"

    print(f"Loading monsters.json …")
    monsters = load_json(monsters_path)
    index = build_monster_index(monsters)
    print(f"  {len(monsters)} monster entries, {len(index)} unique zh_names\n")

    print(f"Loading Excel: {excel_path.name} …")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # ── Write header row for new columns ───────────────────────────────────
    headers = {
        11: "匹配类型 (Match Type)",
        12: "形态数量 (Form Count)",
        13: "形态列表 (Forms in JSON)",
        14: "备注 (Notes)",
    }
    for col, label in headers.items():
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = FONT_BOLD
        cell.alignment = Alignment(wrap_text=True)

    # ── Process each data row ──────────────────────────────────────────────
    counts = defaultdict(int)

    for row_idx in range(2, ws.max_row + 1):
        raw_name = ws.cell(row=row_idx, column=1).value
        if not raw_name:
            continue

        zh_name, form_hint = parse_excel_name(str(raw_name))
        match_type, matched, note = classify_row(zh_name, form_hint, index)
        counts[match_type] += 1

        form_count = len(matched)
        form_list = "; ".join(
            (e["zh_form"] if e["zh_form"] else "default") for e in matched
        ) if matched else "—"

        if dry_run:
            print(f"  Row {row_idx:3d}  [{match_type:<18}]  {raw_name}  →  {note}")
        else:
            fill = FILL[match_type]
            for col, value in [
                (11, match_type),
                (12, form_count if form_count else "—"),
                (13, form_list),
                (14, note),
            ]:
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True)

    # ── Column widths ──────────────────────────────────────────────────────
    if not dry_run:
        ws.column_dimensions[get_column_letter(11)].width = 22
        ws.column_dimensions[get_column_letter(12)].width = 10
        ws.column_dimensions[get_column_letter(13)].width = 45
        ws.column_dimensions[get_column_letter(14)].width = 60

    # ── Summary ────────────────────────────────────────────────────────────
    total = sum(counts.values())
    print(f"\n{'='*60}")
    print(f"Total rows processed : {total}")
    print(f"  EXACT_SINGLE       : {counts['EXACT_SINGLE']}")
    print(f"  EXACT_MULTI_FORM   : {counts['EXACT_MULTI_FORM']}")
    print(f"  PARENTHETICAL      : {counts['PARENTHETICAL']}")
    print(f"  NEW_MONSTER        : {counts['NEW_MONSTER']}")
    print(f"{'='*60}")

    if dry_run:
        print("\n(Dry run — Excel not modified)")
        return

    # Save as a new _marked file (NTFS/WSL cannot overwrite existing files with ACL locks)
    out_path = excel_path.parent / (excel_path.stem + "_marked" + excel_path.suffix)
    wb.save(out_path)
    print(f"\nSaved: {out_path.name}")
    print("(Original file untouched — open the _marked file to review annotations)")


def main():
    dry_run = "--dry-run" in sys.argv or "-n" in sys.argv
    run(dry_run=dry_run)


if __name__ == "__main__":
    main()
