"""
For every EXACT_SINGLE row in data_stats_traits_formal.xlsx, check:

  1. TYPES  — does the Excel type column match main_type / sub_type in monsters.json?
             Excel format: "火" or "火 / 翼"  (spaces around "/" are optional)
             First type  → main_type, second type → sub_type (None if mono-type)

  2. TRAITS — parse the Excel trait cell ("名称：描述") and compare with traits.json:
             a. Trait name change  — monster.trait (English) no longer matches the
                                     English name of the trait named in the Excel cell
             b. Trait description change — traits.json zh description differs from
                                           the (normalised) Excel description
             Unknown trait names (not in traits.json) are reported separately.

A detailed report is written to:
    backend/data/type_trait_check_report_<timestamp>.txt

Usage (from project root):
    python -m backend.scripts.update_stats_traits.check_types_and_traits
"""

import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Punctuation normaliser  (same rules as existing compare_traits.py)
# ---------------------------------------------------------------------------

def normalize_punctuation(text: str) -> str:
    if not text:
        return ""
    text = text.replace(",", "，")
    text = text.replace(";", "；")
    text = text.replace(":", "：")
    text = text.replace("!", "！")
    text = text.replace("?", "？")
    text = text.replace("(", "（")
    text = text.replace(")", "）")
    text = text.replace("[", "【")
    text = text.replace("]", "】")
    text = text.replace(". ", "。")
    text = text.replace(".", "。")
    # remove space after Chinese punctuation
    for p in "，。：；！？）】":
        text = text.replace(p + " ", p)
    if text and not text.endswith("。"):
        text += "。"
    return text


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def parse_excel_name(raw: str):
    """Return (zh_name, form_hint)."""
    raw = raw.strip()
    m = re.match(r'^(.+?)\s*[（(](.+?)[)）]\s*$', raw)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return raw, None


def parse_type_cell(cell_value: str, zh_to_en: dict):
    """
    Parse e.g. "火 / 翼" → ("Fire", "Flying")
    or "草"              → ("Grass", None)

    Returns (main_type_en, sub_type_en) or (None, None) on failure.
    """
    if not cell_value:
        return None, None
    parts = [p.strip() for p in re.split(r"[/／]", str(cell_value))]
    main_zh = parts[0] if len(parts) >= 1 else None
    sub_zh  = parts[1] if len(parts) >= 2 else None
    main_en = zh_to_en.get(main_zh)
    sub_en  = zh_to_en.get(sub_zh) if sub_zh else None
    return main_en, sub_en


def parse_trait_cell(cell_value: str):
    """
    Parse "名称：描述" (Chinese or English colon).
    Returns (trait_name_zh, trait_desc_zh_raw) or (None, None).
    """
    if not cell_value or not isinstance(cell_value, str):
        return None, None
    # Try Chinese colon first, fall back to English colon
    for sep in ("：", ":"):
        if sep in cell_value:
            name, desc = cell_value.split(sep, 1)
            name = name.strip()
            desc = desc.strip()
            if name and desc:
                return name, desc
    return None, None


# ---------------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------------

def load_json(path: Path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def build_type_zh_to_en(types: list) -> dict:
    """{ "草": "Grass", ... }"""
    mapping = {}
    for t in types:
        zh = t.get("localized", {}).get("zh") or t.get("name")
        mapping[zh] = t["name"]
    return mapping


def build_monster_index(monsters: list) -> dict:
    """{ zh_name: [monster_entry, ...] }"""
    idx = defaultdict(list)
    for m in monsters:
        zh = m.get("localized", {}).get("zh", {})
        zh_name = zh.get("name")
        if zh_name:
            idx[zh_name].append(m)
    return idx


def build_trait_zh_index(traits: list) -> dict:
    """{ zh_name: trait_entry }  — keyed by Chinese trait name."""
    idx = {}
    for t in traits:
        zh_name = t.get("localized", {}).get("zh", {}).get("name")
        if zh_name:
            idx[zh_name] = t
    return idx


def build_trait_en_index(traits: list) -> dict:
    """{ en_name: trait_entry }"""
    return {t["name"]: t for t in traits}


# ---------------------------------------------------------------------------
# Classification (mirrors mark_excel.py logic)
# ---------------------------------------------------------------------------

def classify(zh_name: str, form_hint, monster_index: dict) -> str:
    entries = monster_index.get(zh_name)
    if not entries:
        return "NEW_MONSTER"
    if form_hint is not None:
        return "PARENTHETICAL"
    has_real_forms = any(
        e.get("localized", {}).get("zh", {}).get("form") is not None
        for e in entries
    )
    return "EXACT_MULTI_FORM" if has_real_forms else "EXACT_SINGLE"


# ---------------------------------------------------------------------------
# Report builder
# ---------------------------------------------------------------------------

class Report:
    def __init__(self):
        self.type_mismatches     = []   # dict per mismatch
        self.trait_name_changes  = []
        self.trait_desc_changes  = []
        self.trait_unknown       = []   # trait in Excel not in traits.json
        self.ok_count            = 0

    # ── Type helpers ────────────────────────────────────────────────────────

    def add_type_mismatch(self, row, raw_name, en_name, field,
                          excel_val, json_val):
        self.type_mismatches.append({
            "row": row, "raw_name": raw_name, "en_name": en_name,
            "field": field, "excel": excel_val, "json": json_val,
        })

    # ── Trait helpers ────────────────────────────────────────────────────────

    def add_trait_name_change(self, row, raw_name, en_name,
                              old_en_trait, new_zh_trait, new_en_trait):
        self.trait_name_changes.append({
            "row": row, "raw_name": raw_name, "en_name": en_name,
            "old_en_trait": old_en_trait,
            "new_zh_trait": new_zh_trait,
            "new_en_trait": new_en_trait,
        })

    def add_trait_desc_change(self, row, raw_name, en_name,
                              trait_zh_name, trait_en_name,
                              json_desc_norm, excel_desc_norm,
                              json_desc_raw, excel_desc_raw):
        self.trait_desc_changes.append({
            "row": row, "raw_name": raw_name, "en_name": en_name,
            "trait_zh_name": trait_zh_name, "trait_en_name": trait_en_name,
            "json_desc_norm": json_desc_norm, "excel_desc_norm": excel_desc_norm,
            "json_desc_raw": json_desc_raw, "excel_desc_raw": excel_desc_raw,
        })

    def add_trait_unknown(self, row, raw_name, en_name, trait_zh_name, trait_desc_raw):
        self.trait_unknown.append({
            "row": row, "raw_name": raw_name, "en_name": en_name,
            "trait_zh_name": trait_zh_name, "trait_desc_raw": trait_desc_raw,
        })

    # ── Render ──────────────────────────────────────────────────────────────

    def render(self) -> str:
        lines = []
        w = lines.append

        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        w("=" * 80)
        w("TYPE & TRAIT CHECK REPORT")
        w(f"Generated: {ts}")
        w("Scope: EXACT_SINGLE rows in data_stats_traits_formal.xlsx")
        w("=" * 80)
        w("")

        # ── Type mismatches ─────────────────────────────────────────────────
        w(f"{'─'*80}")
        w(f"SECTION 1 — TYPE MISMATCHES  ({len(self.type_mismatches)} found)")
        w(f"{'─'*80}")
        if not self.type_mismatches:
            w("  ✓ All types match.\n")
        else:
            # Group by monster for readability
            by_monster = defaultdict(list)
            for r in self.type_mismatches:
                by_monster[(r["row"], r["raw_name"], r["en_name"])].append(r)
            for (row, raw_name, en_name), records in by_monster.items():
                w(f"\n  Row {row:3d}  {raw_name}  ({en_name})")
                for r in records:
                    w(f"    {r['field']:<12s}  Excel: {str(r['excel']):15s}  JSON: {str(r['json'])}")
            w("")

        # ── Trait name changes ──────────────────────────────────────────────
        w(f"{'─'*80}")
        w(f"SECTION 2 — TRAIT NAME CHANGES  ({len(self.trait_name_changes)} found)")
        w(f"{'─'*80}")
        if not self.trait_name_changes:
            w("  ✓ No trait name changes.\n")
        else:
            for r in self.trait_name_changes:
                w(f"\n  Row {r['row']:3d}  {r['raw_name']}  ({r['en_name']})")
                w(f"    Current trait (EN) : {r['old_en_trait']}")
                w(f"    New trait    (ZH)  : {r['new_zh_trait']}")
                w(f"    New trait    (EN)  : {r['new_en_trait']}")
            w("")

        # ── Trait description changes ───────────────────────────────────────
        w(f"{'─'*80}")
        w(f"SECTION 3 — TRAIT DESCRIPTION CHANGES  ({len(self.trait_desc_changes)} found)")
        w(f"{'─'*80}")
        if not self.trait_desc_changes:
            w("  ✓ No description changes.\n")
        else:
            for r in self.trait_desc_changes:
                w(f"\n  Row {r['row']:3d}  {r['raw_name']}  ({r['en_name']})")
                w(f"  Trait: {r['trait_zh_name']}  ({r['trait_en_name']})")
                w(f"    JSON  : {r['json_desc_norm']}")
                w(f"    Excel : {r['excel_desc_norm']}")
            w("")

        # ── Unknown traits ──────────────────────────────────────────────────
        w(f"{'─'*80}")
        w(f"SECTION 4 — UNKNOWN TRAITS (in Excel but not in traits.json)  ({len(self.trait_unknown)} found)")
        w(f"{'─'*80}")
        if not self.trait_unknown:
            w("  ✓ All trait names recognised.\n")
        else:
            for r in self.trait_unknown:
                w(f"\n  Row {r['row']:3d}  {r['raw_name']}  ({r['en_name']})")
                w(f"    Trait name (ZH) : {r['trait_zh_name']}")
                w(f"    Description     : {r['trait_desc_raw']}")
            w("")

        # ── Summary ─────────────────────────────────────────────────────────
        w(f"{'─'*80}")
        w("SUMMARY")
        w(f"{'─'*80}")
        total_issues = (len(self.type_mismatches) + len(self.trait_name_changes)
                        + len(self.trait_desc_changes) + len(self.trait_unknown))
        w(f"  Rows fully OK             : {self.ok_count}")
        w(f"  Type mismatches           : {len(self.type_mismatches)}")
        w(f"  Trait name changes        : {len(self.trait_name_changes)}")
        w(f"  Trait description changes : {len(self.trait_desc_changes)}")
        w(f"  Unknown traits            : {len(self.trait_unknown)}")
        w(f"  Total issues              : {total_issues}")
        w("")

        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run():
    root = Path(__file__).parent.parent.parent.parent
    data_dir = root / "backend" / "data"

    print("Loading data files …")
    monsters  = load_json(data_dir / "monsters.json")
    types     = load_json(data_dir / "types.json")
    traits    = load_json(data_dir / "traits.json")

    monster_index  = build_monster_index(monsters)
    zh_to_en_type  = build_type_zh_to_en(types)
    trait_zh_idx   = build_trait_zh_index(traits)
    trait_en_idx   = build_trait_en_index(traits)

    print(f"  {len(monsters)} monsters, {len(types)} types, {len(traits)} traits loaded\n")

    print("Loading Excel …")
    wb = openpyxl.load_workbook(data_dir / "data_stats_traits_formal.xlsx", data_only=True)
    ws = wb.active

    report = Report()
    processed = 0

    for row_idx in range(2, ws.max_row + 1):
        raw_name = ws.cell(row=row_idx, column=1).value
        if not raw_name:
            continue

        raw_name = str(raw_name).strip()
        zh_name, form_hint = parse_excel_name(raw_name)

        # Only process EXACT_SINGLE rows
        if classify(zh_name, form_hint, monster_index) != "EXACT_SINGLE":
            continue

        processed += 1
        entries = monster_index[zh_name]
        # For EXACT_SINGLE all entries share the same zh_name with no forms;
        # use the first entry as the representative.
        monster = entries[0]
        en_name = monster.get("name", "?")

        row_has_issue = False

        # ── 1. Type check ─────────────────────────────────────────────────
        type_cell = ws.cell(row=row_idx, column=2).value
        excel_main, excel_sub = parse_type_cell(str(type_cell) if type_cell else "", zh_to_en_type)

        json_main = monster.get("main_type")
        json_sub  = monster.get("sub_type")   # may be None

        if excel_main is None:
            print(f"  ⚠  Row {row_idx}: could not parse type cell '{type_cell}' for {raw_name}")
        else:
            if excel_main != json_main:
                report.add_type_mismatch(row_idx, raw_name, en_name,
                                         "main_type", excel_main, json_main)
                row_has_issue = True

            # Treat None == None as match; treat mismatch otherwise
            if excel_sub != json_sub:
                report.add_type_mismatch(row_idx, raw_name, en_name,
                                         "sub_type",
                                         excel_sub if excel_sub else "(none)",
                                         json_sub  if json_sub  else "(none)")
                row_has_issue = True

        # ── 2. Trait check ───────────────────────────────────────────────
        trait_cell = ws.cell(row=row_idx, column=10).value
        trait_zh_name, trait_desc_raw = parse_trait_cell(str(trait_cell) if trait_cell else "")

        if trait_zh_name is None:
            print(f"  ⚠  Row {row_idx}: could not parse trait cell for {raw_name}: '{trait_cell}'")
        else:
            trait_desc_norm = normalize_punctuation(trait_desc_raw)

            # Look up trait by Chinese name
            trait_entry = trait_zh_idx.get(trait_zh_name)

            if trait_entry is None:
                # Trait name not in traits.json at all
                report.add_trait_unknown(row_idx, raw_name, en_name,
                                         trait_zh_name, trait_desc_raw)
                row_has_issue = True
            else:
                trait_en_name = trait_entry.get("name", "?")
                current_en_trait = monster.get("trait")

                # 2a. Trait name change?
                if current_en_trait != trait_en_name:
                    report.add_trait_name_change(
                        row_idx, raw_name, en_name,
                        current_en_trait, trait_zh_name, trait_en_name
                    )
                    row_has_issue = True

                # 2b. Description change?
                json_desc_raw  = (trait_entry.get("localized", {})
                                             .get("zh", {})
                                             .get("description", ""))
                json_desc_norm = normalize_punctuation(json_desc_raw)

                if json_desc_norm != trait_desc_norm:
                    report.add_trait_desc_change(
                        row_idx, raw_name, en_name,
                        trait_zh_name, trait_en_name,
                        json_desc_norm, trait_desc_norm,
                        json_desc_raw, trait_desc_raw,
                    )
                    row_has_issue = True

        if not row_has_issue:
            report.ok_count += 1

    print(f"  Processed {processed} EXACT_SINGLE rows\n")

    # ── Write report ────────────────────────────────────────────────────────
    report_text = report.render()
    print(report_text)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = data_dir / f"type_trait_check_report_{ts}.txt"
    report_path.write_text(report_text, encoding="utf-8")
    print(f"Report saved: {report_path.name}")


if __name__ == "__main__":
    run()
