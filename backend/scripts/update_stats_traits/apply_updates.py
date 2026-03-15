"""
Apply base stat and trait description updates from data_stats_traits_formal.xlsx
to monsters.json and traits.json.

Scope: EXACT_SINGLE rows only.

STATS (monsters.json):
  - Updates base_hp, base_phy_atk, base_mag_atk, base_phy_def, base_mag_def, base_spd
  - Recalculates preferred_attack_style when atk balance changes

TRAITS (traits.json zh description only — English left untouched):
  - Matched traits  : Excel zh_name found in traits.json → update zh description
  - Unknown traits  : Excel zh_name NOT in traits.json   → update zh description of
                      the monster's CURRENT trait (by English name) — name unchanged

Backups are created before any file is modified:
  monsters.json.backup_<timestamp>
  traits.json.backup_<timestamp>

A change report is written to:
  backend/data/apply_updates_report_<timestamp>.txt

Usage (from project root):
    python -m backend.scripts.update_stats_traits.apply_updates            # interactive
    python -m backend.scripts.update_stats_traits.apply_updates --dry-run  # preview only
"""

import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Punctuation normaliser
# ---------------------------------------------------------------------------

def normalize_punctuation(text: str) -> str:
    if not text:
        return ""
    text = text.replace(",", "，")
    text = text.replace(";", "；")
    text = text.replace("!", "！")
    text = text.replace("?", "？")
    text = text.replace("(", "（")
    text = text.replace(")", "）")
    text = text.replace("[", "【")
    text = text.replace("]", "】")
    text = text.replace(". ", "。")
    text = text.replace(".", "。")
    for p in "，。；！？）】":
        text = text.replace(p + " ", p)
    if text and not text.endswith("。"):
        text += "。"
    return text


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def parse_excel_name(raw: str):
    raw = raw.strip()
    m = re.match(r'^(.+?)\s*[（(](.+?)[)）]\s*$', raw)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return raw, None


def parse_trait_cell(cell_value: str):
    """Return (zh_trait_name, zh_desc_normalized) or (None, None)."""
    if not cell_value:
        return None, None
    for sep in ("：", ":"):
        if sep in cell_value:
            name, desc = cell_value.split(sep, 1)
            name = name.strip()
            desc = desc.strip()
            if name and desc:
                return name, normalize_punctuation(desc)
    return None, None


def preferred_attack_style(phy_atk: int, mag_atk: int) -> str:
    if phy_atk > mag_atk:
        return "Physical"
    if mag_atk > phy_atk:
        return "Magic"
    return "Both"


# ---------------------------------------------------------------------------
# Data loaders / indexes
# ---------------------------------------------------------------------------

def load_json(path: Path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def save_json(path: Path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def backup(path: Path, timestamp: str):
    backup_path = path.parent / f"{path.name}.backup_{timestamp}"
    backup_path.write_bytes(path.read_bytes())
    return backup_path


def build_monster_index(monsters: list) -> dict:
    """{ zh_name: [monster_entry, ...] }"""
    idx = defaultdict(list)
    for m in monsters:
        zh = m.get("localized", {}).get("zh", {})
        zh_name = zh.get("name")
        if zh_name:
            idx[zh_name].append(m)
    return idx


def build_trait_zh_index(traits: list) -> dict:
    """{ zh_name: trait_entry }"""
    return {
        t.get("localized", {}).get("zh", {}).get("name"): t
        for t in traits
        if t.get("localized", {}).get("zh", {}).get("name")
    }


def build_trait_en_index(traits: list) -> dict:
    """{ en_name: trait_entry }"""
    return {t["name"]: t for t in traits}


def classify(zh_name: str, form_hint, monster_index: dict) -> str:
    entries = monster_index.get(zh_name)
    if not entries:
        return "NEW_MONSTER"
    if form_hint is not None:
        return "PARENTHETICAL"
    has_real_forms = any(
        e.get("localized", {}).get("zh", {}).get("form") is not None
        for e in entries
    )
    return "EXACT_MULTI_FORM" if has_real_forms else "EXACT_SINGLE"


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------

class ChangeReport:
    def __init__(self):
        self.stat_changes   = []   # { monster, field, old, new }
        self.trait_changes  = []   # { monster, trait_en, trait_zh, old_desc, new_desc, unknown }
        self.no_stat_change = 0
        self.no_trait_change = 0

    def render(self, dry_run: bool) -> str:
        lines = []
        w = lines.append
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        mode = "DRY RUN — no files modified" if dry_run else "APPLIED"

        w("=" * 80)
        w(f"APPLY UPDATES REPORT  [{mode}]")
        w(f"Generated: {ts}")
        w("=" * 80)
        w("")

        # ── Stats ──────────────────────────────────────────────────────────
        w(f"{'─'*80}")
        w(f"SECTION 1 — BASE STAT CHANGES  ({len(self.stat_changes)} monster(s))")
        w(f"{'─'*80}")
        if not self.stat_changes:
            w("  ✓ No stat changes.\n")
        else:
            STAT_LABELS = {
                "base_hp":      "HP     ",
                "base_phy_atk": "Phy Atk",
                "base_mag_atk": "Mag Atk",
                "base_phy_def": "Phy Def",
                "base_mag_def": "Mag Def",
                "base_spd":     "Speed  ",
                "preferred_attack_style": "Atk Style",
            }
            for entry in self.stat_changes:
                w(f"\n  {entry['zh_name']}  ({entry['en_name']})")
                for chg in entry["changes"]:
                    label = STAT_LABELS.get(chg["field"], chg["field"])
                    w(f"    {label} : {str(chg['old']):>12}  →  {chg['new']}")
            w("")

        # ── Trait descriptions ─────────────────────────────────────────────
        w(f"{'─'*80}")
        w(f"SECTION 2 — TRAIT DESCRIPTION CHANGES  ({len(self.trait_changes)} trait(s))")
        w(f"{'─'*80}")
        if not self.trait_changes:
            w("  ✓ No trait description changes.\n")
        else:
            for entry in self.trait_changes:
                unknown_note = "  ⚠ UNKNOWN TRAIT NAME IN EXCEL — used current trait" if entry["unknown"] else ""
                w(f"\n  Trait: {entry['trait_zh']}  ({entry['trait_en']}){unknown_note}")
                w(f"  Used by: {', '.join(entry['monsters'])}")
                w(f"    Old : {entry['old_desc']}")
                w(f"    New : {entry['new_desc']}")
            w("")

        # ── Summary ────────────────────────────────────────────────────────
        w(f"{'─'*80}")
        w("SUMMARY")
        w(f"{'─'*80}")
        w(f"  Monsters with stat changes     : {len(self.stat_changes)}")
        w(f"  Monsters with no stat change   : {self.no_stat_change}")
        w(f"  Traits with description change : {len(self.trait_changes)}")
        w(f"  Traits with no change          : {self.no_trait_change}")
        w("")

        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Core logic
# ---------------------------------------------------------------------------

STAT_FIELDS = [
    "base_hp", "base_phy_atk", "base_mag_atk",
    "base_phy_def", "base_mag_def", "base_spd",
]


def run(dry_run: bool = False):
    root = Path(__file__).parent.parent.parent.parent
    data_dir = root / "backend" / "data"

    print("Loading data files …")
    monsters = load_json(data_dir / "monsters.json")
    traits   = load_json(data_dir / "traits.json")

    monster_index  = build_monster_index(monsters)
    trait_zh_idx   = build_trait_zh_index(traits)
    trait_en_idx   = build_trait_en_index(traits)

    print(f"  {len(monsters)} monsters, {len(traits)} traits\n")

    print("Loading Excel …")
    wb = openpyxl.load_workbook(data_dir / "data_stats_traits_formal.xlsx", data_only=True)
    ws = wb.active

    report = ChangeReport()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Track which trait entries need description updates:
    # { trait_en_name: (new_zh_desc, [monster_zh_names], is_unknown) }
    trait_updates: dict[str, tuple[str, list[str], bool]] = {}

    # ── Pass 1: collect all changes ─────────────────────────────────────────
    for row_idx in range(2, ws.max_row + 1):
        raw_name = ws.cell(row=row_idx, column=1).value
        if not raw_name:
            continue

        raw_name = str(raw_name).strip()
        zh_name, form_hint = parse_excel_name(raw_name)

        if classify(zh_name, form_hint, monster_index) != "EXACT_SINGLE":
            continue

        entries = monster_index[zh_name]
        representative = entries[0]
        en_name = representative.get("name", "?")

        # ── Stats ───────────────────────────────────────────────────────────
        excel_stats = {
            "base_hp":      ws.cell(row=row_idx, column=4).value,
            "base_phy_atk": ws.cell(row=row_idx, column=5).value,
            "base_mag_atk": ws.cell(row=row_idx, column=6).value,
            "base_phy_def": ws.cell(row=row_idx, column=7).value,
            "base_mag_def": ws.cell(row=row_idx, column=8).value,
            "base_spd":     ws.cell(row=row_idx, column=9).value,
        }

        # Validate all stat values are present
        if any(v is None for v in excel_stats.values()):
            print(f"  ⚠  Row {row_idx}: missing stat value(s) for {raw_name} — skipped")
            continue

        stat_entry = {"zh_name": zh_name, "en_name": en_name, "changes": []}

        for monster in entries:
            for field in STAT_FIELDS:
                old_val = monster.get(field)
                new_val = excel_stats[field]
                if old_val != new_val:
                    stat_entry["changes"].append({"field": field, "old": old_val, "new": new_val})

            # Check preferred_attack_style
            new_phy = excel_stats["base_phy_atk"]
            new_mag = excel_stats["base_mag_atk"]
            new_style = preferred_attack_style(new_phy, new_mag)
            old_style = monster.get("preferred_attack_style")
            if old_style != new_style:
                stat_entry["changes"].append({
                    "field": "preferred_attack_style",
                    "old": old_style,
                    "new": new_style,
                })

        if stat_entry["changes"]:
            report.stat_changes.append(stat_entry)
        else:
            report.no_stat_change += 1

        # ── Traits ──────────────────────────────────────────────────────────
        trait_cell = ws.cell(row=row_idx, column=10).value
        excel_trait_zh, excel_desc_norm = parse_trait_cell(
            str(trait_cell) if trait_cell else ""
        )

        if excel_trait_zh is None:
            print(f"  ⚠  Row {row_idx}: could not parse trait cell for {raw_name}")
            continue

        is_unknown = excel_trait_zh not in trait_zh_idx

        if is_unknown:
            # Use the monster's current trait (by English name)
            current_en_trait = representative.get("trait")
            trait_entry = trait_en_idx.get(current_en_trait)
            if trait_entry is None:
                print(f"  ⚠  Row {row_idx}: current trait '{current_en_trait}' not found in traits.json — skipped")
                continue
            trait_en_name = current_en_trait
            trait_zh_name = trait_entry.get("localized", {}).get("zh", {}).get("name", "?")
        else:
            trait_entry = trait_zh_idx[excel_trait_zh]
            trait_en_name = trait_entry.get("name", "?")
            trait_zh_name = excel_trait_zh

        current_desc = trait_entry.get("localized", {}).get("zh", {}).get("description", "")
        current_desc_norm = normalize_punctuation(current_desc)

        if current_desc_norm == excel_desc_norm:
            report.no_trait_change += 1
            continue

        # Accumulate: multiple monsters may share the same trait
        if trait_en_name not in trait_updates:
            trait_updates[trait_en_name] = (excel_desc_norm, [], is_unknown)
        else:
            # Warn if two monsters give conflicting descriptions for the same trait
            existing_desc = trait_updates[trait_en_name][0]
            if existing_desc != excel_desc_norm:
                print(f"  ⚠  Trait '{trait_en_name}': conflicting descriptions from different monsters "
                      f"('{zh_name}' vs earlier row) — keeping first occurrence")
        trait_updates[trait_en_name][1].append(zh_name)

    # De-duplicate trait report entries and add to report
    for trait_en_name, (new_desc, monster_list, is_unknown) in trait_updates.items():
        trait_entry = trait_en_idx[trait_en_name]
        trait_zh_name = trait_entry.get("localized", {}).get("zh", {}).get("name", "?")
        old_desc = normalize_punctuation(
            trait_entry.get("localized", {}).get("zh", {}).get("description", "")
        )
        report.trait_changes.append({
            "trait_en":  trait_en_name,
            "trait_zh":  trait_zh_name,
            "old_desc":  old_desc,
            "new_desc":  new_desc,
            "monsters":  monster_list,
            "unknown":   is_unknown,
        })

    # ── Preview ─────────────────────────────────────────────────────────────
    report_text = report.render(dry_run)
    print(report_text)

    if dry_run:
        print("(Dry run — no files modified)")
        return

    if not report.stat_changes and not report.trait_changes:
        print("Nothing to update.")
        return

    # ── Confirm ─────────────────────────────────────────────────────────────
    response = input("Apply these changes? (yes/no): ").strip().lower()
    if response not in ("yes", "y"):
        print("Aborted — no changes made.")
        return

    # ── Apply stat changes ───────────────────────────────────────────────────
    if report.stat_changes:
        bp = backup(data_dir / "monsters.json", timestamp)
        print(f"\nBackup: {bp.name}")

        # Re-index so we can update by zh_name
        changed_names = {e["zh_name"] for e in report.stat_changes}

        # Re-read Excel stats into a quick lookup
        excel_stat_lookup: dict[str, dict] = {}
        for row_idx in range(2, ws.max_row + 1):
            raw_name = ws.cell(row=row_idx, column=1).value
            if not raw_name:
                continue
            zh_name, form_hint = parse_excel_name(str(raw_name).strip())
            if classify(zh_name, form_hint, monster_index) != "EXACT_SINGLE":
                continue
            excel_stat_lookup[zh_name] = {
                "base_hp":      ws.cell(row=row_idx, column=4).value,
                "base_phy_atk": ws.cell(row=row_idx, column=5).value,
                "base_mag_atk": ws.cell(row=row_idx, column=6).value,
                "base_phy_def": ws.cell(row=row_idx, column=7).value,
                "base_mag_def": ws.cell(row=row_idx, column=8).value,
                "base_spd":     ws.cell(row=row_idx, column=9).value,
            }

        for monster in monsters:
            zh_name = monster.get("localized", {}).get("zh", {}).get("name")
            if zh_name not in changed_names:
                continue
            stats = excel_stat_lookup.get(zh_name)
            if not stats:
                continue
            for field in STAT_FIELDS:
                monster[field] = stats[field]
            monster["preferred_attack_style"] = preferred_attack_style(
                stats["base_phy_atk"], stats["base_mag_atk"]
            )

        save_json(data_dir / "monsters.json", monsters)
        print("monsters.json updated.")

    # ── Apply trait description changes ──────────────────────────────────────
    if report.trait_changes:
        bp = backup(data_dir / "traits.json", timestamp)
        print(f"Backup: {bp.name}")

        for entry in report.trait_changes:
            trait_entry = trait_en_idx[entry["trait_en"]]
            trait_entry["localized"]["zh"]["description"] = entry["new_desc"]

        save_json(data_dir / "traits.json", traits)
        print("traits.json updated.")

    # ── Save report ──────────────────────────────────────────────────────────
    report_path = data_dir / f"apply_updates_report_{timestamp}.txt"
    report_path.write_text(report_text, encoding="utf-8")
    print(f"\nReport saved: {report_path.name}")
    print("\nDone. Run check_local_consistency.py to validate, then update_game_data.py to push to DB.")


def main():
    dry_run = "--dry-run" in sys.argv or "-n" in sys.argv
    run(dry_run)


if __name__ == "__main__":
    main()
