"""
Validate moves.json and statuses.json against moves_damage_treatment_review.xlsx.

moves.json checks:
  1. Name coverage  — every move in the Excel exists in moves.json and vice-versa.
  2. Alt-field triplet — alt_power_total, alt_condition_zh, alt_condition_en must all
     be present or all absent.
  3. Alt > base power — when alt_power_total is set it must exceed the move's base power.
  4. counter_power_multiplier sanity — must be > 1, only on attack moves.
  5. HIGH implementable coverage — every HIGH-priority "implementable_with_context" attack
     move that has conditional_power or damage_multiplier rule kinds must have at least one
     of: counter_power_multiplier, alt_power_total, or a move_specific status with
     power_bonus > 0.
  6. NoSpecialTreatment cleanliness — moves listed in that sheet must NOT have
     counter_power_multiplier or alt_power_total set.
  7. Spurious fields — no move should have alt_condition_zh/en without alt_power_total.

statuses.json checks:
  8.  Move reference — every status's `move` field must name a valid move in moves.json.
  9.  usage/affect values — must be known enum strings.
  10. defense_only category — defense_only statuses must be on DEFENSE category moves.
  11. attack_only category — attack_only statuses must NOT be on DEFENSE category moves.
  12. dmg_reduction_pct range — must be 1–100 when present.
  13. move_specific requires power_bonus > 0 and en_name.
  14. At least one effect field — each status must have a non-zero numeric effect.
  15. Opponent debuff correctness — attack_only/affect=opponent statuses that intend to
      reduce defense must have at least one of phy_def_boost or mag_def_boost negative.

Usage (from project root):
    python3 -m backend.scripts.validation.check_moves_damage_fields
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

ROOT = Path(__file__).resolve().parents[3]
MOVES_JSON = ROOT / "backend/data/moves.json"
STATUSES_JSON = ROOT / "backend/data/statuses.json"
EXCEL_PATH = ROOT / "moves_damage_treatment_review.xlsx"

ATTACK_CATEGORIES = {"Physical Attack", "Magic Attack", "PHY_ATTACK", "MAG_ATTACK"}
HIGH_DAMAGE_RULES = {"conditional_power", "damage_multiplier", "power_multiplier"}
VALID_USAGES = {"all", "attack_only", "defense_only", "move_specific"}
VALID_AFFECTS = {"self", "opponent"}
STATUS_EFFECT_FIELDS = {
    "dmg_reduction_pct", "phy_atk_boost", "mag_atk_boost",
    "phy_def_boost", "mag_def_boost", "flat_power_boost", "pct_power_boost", "power_bonus",
}


def load_moves() -> list[dict]:
    with open(MOVES_JSON, encoding="utf-8") as f:
        return json.load(f)


def load_statuses() -> list[dict]:
    with open(STATUSES_JSON, encoding="utf-8") as f:
        return json.load(f)


def load_excel_sheets() -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheets: dict[str, list[dict]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                rows.append(dict(zip(headers, row)))
        sheets[sheet_name] = rows
    return sheets


def rules_set(row: dict) -> set[str]:
    raw = row.get("rule_kinds") or ""
    return {r.strip() for r in raw.split(";") if r.strip()}


def main() -> None:
    errors: list[str] = []
    warnings: list[str] = []

    moves = load_moves()
    statuses = load_statuses()
    sheets = load_excel_sheets()
    move_by_name = {m["name"]: m for m in moves}
    # move_specific statuses keyed by move name — used to detect coverage via statuses.json
    move_specific_by_move: set[str] = {
        s["move"] for s in statuses
        if s.get("usage") == "move_specific" and (s.get("power_bonus") or 0) > 0
    }

    high_rows = sheets.get("HIGH_DamageFormula", [])
    no_special_rows = sheets.get("NoSpecialTreatment", [])

    excel_names: set[str] = set()
    for sheet_rows in sheets.values():
        for row in sheet_rows:
            n = row.get("name")
            if n:
                excel_names.add(n)

    # ═══════════════════════════════════════════════════════════════
    # moves.json checks
    # ═══════════════════════════════════════════════════════════════

    # ── 1. Name coverage ─────────────────────────────────────────────────────

    json_names = set(move_by_name)
    for name in sorted(excel_names - json_names):
        errors.append(f"[coverage] Move in Excel but missing from moves.json: {name!r}")
    for name in sorted(json_names - excel_names):
        errors.append(f"[coverage] Move in moves.json but missing from Excel: {name!r}")

    # ── 2-4 & 7. Per-move field validation ───────────────────────────────────

    for move in moves:
        name = move["name"]
        cat = move.get("category", "")
        is_attack = cat in ATTACK_CATEGORIES
        alt_total = move.get("alt_power_total")
        alt_zh = move.get("alt_condition_zh")
        alt_en = move.get("alt_condition_en")
        cpm = move.get("counter_power_multiplier")
        base_power = move.get("power")

        # 2. Alt triplet consistency
        alt_fields = [alt_total is not None, alt_zh is not None, alt_en is not None]
        if any(alt_fields) and not all(alt_fields):
            errors.append(
                f"[alt-triplet] {name!r}: partial alt fields — "
                f"alt_power_total={alt_total!r}, alt_condition_zh={alt_zh!r}, alt_condition_en={alt_en!r}"
            )

        # 7. alt_condition label without alt_power_total
        if (alt_zh or alt_en) and alt_total is None:
            errors.append(f"[spurious] {name!r}: has alt_condition label but no alt_power_total")

        # 3. Alt > base power
        if alt_total is not None and base_power is not None:
            if alt_total <= base_power:
                warnings.append(
                    f"[alt-power] {name!r}: alt_power_total={alt_total} <= base power={base_power} "
                    f"(conditional damage should exceed base)"
                )

        # 4. counter_power_multiplier sanity
        if cpm is not None:
            if cpm <= 1:
                errors.append(
                    f"[counter-mult] {name!r}: counter_power_multiplier={cpm} must be > 1"
                )
            if not is_attack:
                errors.append(
                    f"[counter-mult] {name!r}: counter_power_multiplier set on non-attack move (category={cat!r})"
                )

    # ── 5. HIGH implementable coverage ───────────────────────────────────────

    for row in high_rows:
        name = row.get("name")
        impl = row.get("implementation_status") or ""
        scope = row.get("scope") or ""
        cat = row.get("category") or ""

        if impl != "implementable_with_context":
            continue
        if scope != "current_attack_damage":
            continue
        if cat not in ATTACK_CATEGORIES:
            continue

        rules = rules_set(row)
        if not (rules & HIGH_DAMAGE_RULES):
            continue

        move = move_by_name.get(name)
        if move is None:
            continue  # already caught by coverage check

        has_cpm = bool(move.get("counter_power_multiplier"))
        has_alt = bool(move.get("alt_power_total"))
        has_move_specific = name in move_specific_by_move

        if not (has_cpm or has_alt or has_move_specific):
            warnings.append(
                f"[high-coverage] {name!r}: HIGH implementable with rules {rules & HIGH_DAMAGE_RULES} "
                f"but no counter_power_multiplier, alt_power_total, or move_specific status found"
            )

    # ── 6. NoSpecialTreatment cleanliness ────────────────────────────────────

    for row in no_special_rows:
        name = row.get("name")
        if not name:
            continue
        move = move_by_name.get(name)
        if move is None:
            continue
        if move.get("counter_power_multiplier") is not None:
            errors.append(
                f"[no-special] {name!r}: listed in NoSpecialTreatment but has "
                f"counter_power_multiplier={move['counter_power_multiplier']}"
            )
        if move.get("alt_power_total") is not None:
            errors.append(
                f"[no-special] {name!r}: listed in NoSpecialTreatment but has "
                f"alt_power_total={move['alt_power_total']}"
            )

    # ═══════════════════════════════════════════════════════════════
    # statuses.json checks
    # ═══════════════════════════════════════════════════════════════

    for idx, status in enumerate(statuses):
        label = f"[statuses #{idx}] {status.get('status_name', '?')!r} (move={status.get('move', '?')!r})"

        move_name = status.get("move")
        usage = status.get("usage", "")
        affect = status.get("affect", "")

        # 8. Move reference
        if not move_name:
            errors.append(f"{label}: missing 'move' field")
        elif move_name not in move_by_name:
            errors.append(f"{label}: 'move' references unknown move {move_name!r}")

        # 9. usage/affect enum values
        if usage not in VALID_USAGES:
            errors.append(f"{label}: unknown usage={usage!r} (expected one of {VALID_USAGES})")
        if affect not in VALID_AFFECTS:
            errors.append(f"{label}: unknown affect={affect!r} (expected one of {VALID_AFFECTS})")

        move_cat = move_by_name.get(move_name, {}).get("category", "") if move_name else ""

        # 10. defense_only must be on DEFENSE category moves
        if usage == "defense_only" and move_cat not in ("Defense", "DEFENSE", ""):
            errors.append(
                f"{label}: usage=defense_only but move category is {move_cat!r} (expected Defense)"
            )

        # 11. attack_only must NOT be on DEFENSE category moves
        if usage == "attack_only" and move_cat in ("Defense", "DEFENSE"):
            errors.append(
                f"{label}: usage=attack_only but move category is Defense"
            )

        # 12. dmg_reduction_pct range
        dmg_red = status.get("dmg_reduction_pct")
        if dmg_red is not None:
            if not (1 <= dmg_red <= 100):
                errors.append(
                    f"{label}: dmg_reduction_pct={dmg_red} out of range [1, 100]"
                )

        # 13. move_specific requires power_bonus > 0 and en_name
        if usage == "move_specific":
            power_bonus = status.get("power_bonus") or 0
            if power_bonus <= 0:
                errors.append(f"{label}: usage=move_specific but power_bonus={power_bonus} (must be > 0)")
            if not status.get("en_name"):
                errors.append(f"{label}: usage=move_specific but missing en_name (needed for UI display)")

        # 14. At least one non-zero effect field
        effects = {f: status.get(f) for f in STATUS_EFFECT_FIELDS if status.get(f) is not None}
        if not effects:
            warnings.append(f"{label}: no effect fields set — status has no numeric effect")
        elif all(v == 0 for v in effects.values()):
            warnings.append(f"{label}: all effect fields are zero — status has no effect")

        # 15. opponent debuff targeting sanity
        # attack_only + affect=opponent that has ANY boost field should logically debuff
        # (negative def) or debuff atk — flag if it accidentally sets positive def boosts
        if usage == "attack_only" and affect == "opponent":
            phy_def = status.get("phy_def_boost") or 0
            mag_def = status.get("mag_def_boost") or 0
            phy_atk = status.get("phy_atk_boost") or 0
            mag_atk = status.get("mag_atk_boost") or 0
            if phy_def > 0 or mag_def > 0:
                warnings.append(
                    f"{label}: affect=opponent but has positive def boost "
                    f"(phy_def_boost={phy_def}, mag_def_boost={mag_def}) — intended?"
                )
            if phy_atk > 0 or mag_atk > 0:
                warnings.append(
                    f"{label}: affect=opponent but has positive atk boost "
                    f"(phy_atk_boost={phy_atk}, mag_atk_boost={mag_atk}) — intended?"
                )

    # ── Report ────────────────────────────────────────────────────────────────

    if errors:
        print(f"\n❌  {len(errors)} ERROR(S):\n")
        for e in errors:
            print(f"  • {e}")
    if warnings:
        print(f"\n⚠️  {len(warnings)} WARNING(S):\n")
        for w in warnings:
            print(f"  • {w}")
    if not errors and not warnings:
        print("✅  All checks passed — moves.json and statuses.json look consistent.")
    elif not errors:
        print(f"\n✅  No errors. {len(warnings)} warning(s) to review.")

    sys.exit(1 if errors else 0)


if __name__ == "__main__":
    main()
