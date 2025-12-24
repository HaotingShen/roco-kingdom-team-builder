#!/usr/bin/env python3
"""
Source Data Correctness Checker for Roco Kingdom Team Builder

This script validates local JSON data files against the source Excel file (data_all.xlsx).
It checks:
- Monster data correctness (stats, traits, types, names)
- Move data correctness (moveset consistency)
- Data completeness (missing or extra entries)
"""

import json
import sys
from pathlib import Path
from typing import Dict, List, Set, Optional, Tuple
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


class Colors:
    """ANSI color codes for terminal output."""
    RED = '\033[91m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    MAGENTA = '\033[95m'
    CYAN = '\033[96m'
    RESET = '\033[0m'
    BOLD = '\033[1m'


class SourceCorrectnessChecker:
    def __init__(self, data_dir: Path, excel_path: Path):
        self.data_dir = data_dir
        self.excel_path = excel_path
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.info: List[str] = []

        # Load local JSON files
        self.monsters = self._load_json("monsters.json")
        self.moves = self._load_json("moves.json")
        self.monster_moves = self._load_json("monster_moves.json")
        self.traits = self._load_json("traits.json")
        self.types = self._load_json("types.json")

        # Build trait and type lookup maps
        self.trait_zh_to_en = {t["localized"]["zh"]["name"]: t["name"] for t in self.traits}
        self.type_zh_to_en = {t["localized"]["zh"]: t["name"] for t in self.types}

        # Build move name maps
        self.move_zh_to_en = {}
        for move in self.moves:
            zh_name = move.get("localized", {}).get("zh", {}).get("name")
            if zh_name:
                self.move_zh_to_en[zh_name] = move["name"]

        # Build JSON monster lookup by (zh_base_name, zh_form) composite key
        self.json_monster_map = {}
        for monster in self.monsters:
            zh_base = monster.get("localized", {}).get("zh", {}).get("name")
            zh_form = monster.get("localized", {}).get("zh", {}).get("form")

            if zh_base:
                # Key is (base_name, form) or (base_name, None) for default forms
                key = (zh_base, zh_form)
                self.json_monster_map[key] = monster

        # Load Excel data
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.excel_monsters = self._load_excel_monsters()
        self.excel_moves = self._load_excel_moves()

    def _load_json(self, filename: str) -> any:
        """Load and parse a JSON file."""
        filepath = self.data_dir / filename
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            self.error(f"File not found: {filename}")
            sys.exit(1)
        except json.JSONDecodeError as e:
            self.error(f"Invalid JSON in {filename}: {e}")
            sys.exit(1)

    def _load_excel_monsters(self) -> List[dict]:
        """Load monster data from Excel file (in order)."""
        ws = self.wb['基础属性']
        monsters = []

        # Read headers
        headers = [cell.value for cell in ws[1]]

        # First pass: identify leader monsters with multiple forms
        # A leader has multiple forms if it appears multiple times consecutively
        multi_form_leaders = set()
        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]
            monster_name = row[1].value
            stage = row[2].value

            if not monster_name or stage != "首领形态":
                continue

            # Check if next row(s) have the same leader monster name
            for next_row_idx in range(row_idx + 1, min(row_idx + 10, ws.max_row + 1)):
                next_row = ws[next_row_idx]
                next_name = next_row[1].value
                next_stage = next_row[2].value

                if next_name == monster_name and next_stage == "首领形态":
                    # Found another instance of the same leader → multi-form leader
                    multi_form_leaders.add(monster_name)
                    break
                elif next_stage == "首领形态" and next_name != monster_name:
                    # Different leader encountered, stop searching
                    break

        # Track previous row to determine leader forms
        prev_monster_name = None

        # Second pass: process each row
        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]

            # Get monster name (column 2)
            monster_name = row[1].value
            if not monster_name:
                continue

            stage = row[2].value  # Column 3

            # For leader monsters with multiple forms, append form from previous evolution stage
            # Only apply this logic if:
            # 1. Current monster is a leader (stage == "首领形态")
            # 2. Leader has multiple forms (appears multiple times)
            # 3. Previous row had a monster name
            full_name = monster_name
            if (stage == "首领形态" and
                monster_name in multi_form_leaders and
                prev_monster_name):

                # Special case: 棋契陛下 uses the FULL previous monster name as form
                # (e.g., "棋齐垒-黑子" instead of just "黑子")
                if monster_name == "棋契陛下":
                    full_name = f"{monster_name}-{prev_monster_name}"
                elif "-" in prev_monster_name:
                    # Normal case: extract just the form suffix (everything after the first '-')
                    form_part = prev_monster_name.split("-", 1)[1]
                    full_name = f"{monster_name}-{form_part}"

            # Parse type (column 5-6)
            # Column 5: main_type, Column 6: sub_type (separate cells)
            main_type_value = row[4].value  # Column 5
            sub_type_value = row[5].value   # Column 6

            main_type = str(main_type_value).strip() if main_type_value else None
            sub_type = str(sub_type_value).strip() if sub_type_value else None

            # Parse trait (column 7) - extract just the name before ":"
            trait_full = row[6].value  # Column 7
            trait_name = None
            if trait_full:
                trait_str = str(trait_full).strip()
                # Try Chinese colon first, then English colon
                if "：" in trait_str:
                    trait_name = trait_str.split("：")[0].strip()
                elif ":" in trait_str:
                    trait_name = trait_str.split(":")[0].strip()
                elif trait_str:
                    # If no colon found, use the whole string (might be just trait name)
                    trait_name = trait_str

            monster_data = {
                "id": row[0].value,  # Column 1
                "name": full_name,  # Use full_name which includes form for leaders
                "original_name": monster_name,  # Keep original name for reference
                "stage": stage,
                "evolution_condition": row[3].value,  # Column 4
                "main_type": main_type,
                "sub_type": sub_type,
                "trait": trait_name,
                "base_hp": row[7].value,  # Column 8
                "base_phy_atk": row[8].value,  # Column 9
                "base_mag_atk": row[9].value,  # Column 10
                "base_phy_def": row[10].value,  # Column 11
                "base_mag_def": row[11].value,  # Column 12
                "base_spd": row[12].value,  # Column 13
            }

            monsters.append(monster_data)

            # Update previous monster name for next iteration
            prev_monster_name = monster_name

        return monsters

    def _load_excel_moves(self) -> Dict[str, dict]:
        """Load move data from Excel file."""
        ws = self.wb['技能表']
        moves_data = {}

        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]

            monster_name = row[1].value  # Column 2
            if not monster_name:
                continue

            # Parse learnable moves (columns 3-40)
            # Format: "1 撞击" where first part is level, second is move name
            # Some cells may contain only level number (e.g., "1", "10") - skip these
            learnable = []
            for col_idx in range(2, 40):  # Columns 3-40
                move = row[col_idx].value
                if move and move != "-" and str(move).strip():
                    move_str = str(move).strip()
                    # Check if format is "level movename" (e.g., "1 撞击")
                    parts = move_str.split(None, 1)  # Split on whitespace, max 2 parts
                    if len(parts) == 2 and parts[0].isdigit():
                        # Extract just the move name (second part)
                        learnable.append(parts[1].strip())
                    elif len(parts) == 1 and parts[0].isdigit():
                        # Only a number, no move name - skip this cell
                        continue
                    else:
                        # No level prefix, use as-is
                        learnable.append(move_str)

            # Parse move stones (columns 41-61)
            move_stones = []
            for col_idx in range(40, 61):  # Columns 41-61
                move = row[col_idx].value
                if move and move != "-" and str(move).strip():
                    move_stones.append(str(move).strip())

            # Parse legacy moves (columns 62-79)
            legacy = []
            for col_idx in range(61, 79):  # Columns 62-79
                move = row[col_idx].value
                if move and move != "-" and str(move).strip():
                    legacy.append(str(move).strip())

            moves_data[monster_name] = {
                "learnable_moves": learnable,
                "move_stones": move_stones,
                "legacy_moves": legacy,
            }

        return moves_data

    def error(self, msg: str):
        """Record an error message."""
        self.errors.append(msg)

    def warning(self, msg: str):
        """Record a warning message."""
        self.warnings.append(msg)

    def info_msg(self, msg: str):
        """Record an info message."""
        self.info.append(msg)

    def check_all(self):
        """Run all correctness checks."""
        print(f"{Colors.BOLD}{Colors.CYAN}Starting Source Data Correctness Checks...{Colors.RESET}\n")
        print(f"Source: {self.excel_path}\n")

        self.check_monster_coverage()
        self.check_monster_stats()
        self.check_monster_types()
        self.check_monster_traits()
        self.check_monster_forms()
        self.check_evolution_chains()
        self.check_preferred_attack_style()
        self.check_moveset_keys()
        self.check_move_coverage()
        self.check_movesets()

        return self.print_summary()

    def check_monster_coverage(self):
        """Check if all monsters from Excel are in JSON and vice versa."""
        print(f"{Colors.BOLD}Checking monster data coverage...{Colors.RESET}")

        # Build full monster identifiers from Excel (base name + form if exists)
        excel_identifiers = []
        for excel_data in self.excel_monsters:
            # excel_data["name"] is the full name including form (e.g., "丢丢-沙地附近的样子")
            excel_identifiers.append(excel_data["name"])

        # Build full monster identifiers from JSON (base name + form if exists)
        json_identifiers = []
        for monster in self.monsters:
            zh_base = monster.get("localized", {}).get("zh", {}).get("name")
            zh_form = monster.get("localized", {}).get("zh", {}).get("form")

            if zh_base:
                if zh_form:
                    # Has form: "base-form"
                    json_identifiers.append(f"{zh_base}-{zh_form}")
                else:
                    # No form: just base name
                    json_identifiers.append(zh_base)

        excel_set = set(excel_identifiers)
        json_set = set(json_identifiers)

        # Find monsters in Excel but not in JSON
        missing_in_json = excel_set - json_set
        if missing_in_json:
            for name in sorted(missing_in_json):
                self.error(f"Monster '{name}' exists in Excel but not in monsters.json")

        # Find monsters in JSON but not in Excel
        extra_in_json = json_set - excel_set
        if extra_in_json:
            for name in sorted(extra_in_json):
                self.info_msg(f"Monster '{name}' exists in monsters.json but not in Excel (will be skipped)")

        print(f"  Excel monsters: {len(excel_identifiers)}")
        print(f"  JSON monsters: {len(json_identifiers)}")
        print(f"  Matched monsters: {min(len(excel_identifiers), len(json_identifiers))}")
        print(f"  ✓ Coverage check complete\n")

    def check_monster_stats(self):
        """Check if monster base stats match between Excel and JSON (name-based matching)."""
        print(f"{Colors.BOLD}Checking monster base stats...{Colors.RESET}")

        stat_mismatches = 0
        stat_fields = ["base_hp", "base_phy_atk", "base_mag_atk",
                       "base_phy_def", "base_mag_def", "base_spd"]

        # Use Excel as base, find matching JSON monster by name
        for i, excel_data in enumerate(self.excel_monsters):
            excel_full_name = excel_data["name"]

            # Parse Excel name to get base + form
            if "-" in excel_full_name:
                parts = excel_full_name.split("-", 1)
                zh_base = parts[0].strip()
                zh_form = parts[1].strip()
            else:
                zh_base = excel_full_name
                zh_form = None

            # Look up matching JSON monster
            json_monster = self.json_monster_map.get((zh_base, zh_form))

            if not json_monster:
                # Monster not found - already reported in coverage check
                continue

            monster_id = f"{json_monster['name']} ({excel_full_name}) [Excel row {i+2}]"

            for stat in stat_fields:
                json_value = json_monster.get(stat)
                excel_value = excel_data.get(stat)

                if json_value != excel_value:
                    self.error(f"{monster_id}: {stat} mismatch - JSON: {json_value}, Excel: {excel_value}")
                    stat_mismatches += 1

        if stat_mismatches == 0:
            print(f"  ✓ All base stats match\n")
        else:
            print(f"  Found {stat_mismatches} stat mismatches\n")

    def check_monster_types(self):
        """Check if monster types match between Excel and JSON (name-based matching)."""
        print(f"{Colors.BOLD}Checking monster types...{Colors.RESET}")

        type_mismatches = 0

        for i, excel_data in enumerate(self.excel_monsters):
            excel_full_name = excel_data["name"]

            # Parse Excel name
            if "-" in excel_full_name:
                parts = excel_full_name.split("-", 1)
                zh_base = parts[0].strip()
                zh_form = parts[1].strip()
            else:
                zh_base = excel_full_name
                zh_form = None

            json_monster = self.json_monster_map.get((zh_base, zh_form))
            if not json_monster:
                continue

            monster_id = f"{json_monster['name']} ({excel_full_name}) [Excel row {i+2}]"

            # Convert Excel Chinese types to English
            excel_main_type_en = self.type_zh_to_en.get(excel_data["main_type"])
            excel_sub_type_en = self.type_zh_to_en.get(excel_data["sub_type"]) if excel_data["sub_type"] else None

            json_main = json_monster.get("main_type")
            json_sub = json_monster.get("sub_type")

            if json_main != excel_main_type_en:
                self.error(f"{monster_id}: main_type mismatch - JSON: {json_main}, Excel: {excel_main_type_en} ({excel_data['main_type']})")
                type_mismatches += 1

            if json_sub != excel_sub_type_en:
                self.error(f"{monster_id}: sub_type mismatch - JSON: {json_sub}, Excel: {excel_sub_type_en}")
                type_mismatches += 1

        if type_mismatches == 0:
            print(f"  ✓ All types match\n")
        else:
            print(f"  Found {type_mismatches} type mismatches\n")

    def check_monster_traits(self):
        """Check if monster traits match between Excel and JSON (name-based matching)."""
        print(f"{Colors.BOLD}Checking monster traits...{Colors.RESET}")

        trait_mismatches = 0

        for i, excel_data in enumerate(self.excel_monsters):
            excel_full_name = excel_data["name"]

            # Parse Excel name
            if "-" in excel_full_name:
                parts = excel_full_name.split("-", 1)
                zh_base = parts[0].strip()
                zh_form = parts[1].strip()
            else:
                zh_base = excel_full_name
                zh_form = None

            json_monster = self.json_monster_map.get((zh_base, zh_form))
            if not json_monster:
                continue

            monster_id = f"{json_monster['name']} ({excel_full_name}) [Excel row {i+2}]"

            excel_trait = excel_data.get("trait")
            if not excel_trait:
                continue

            # Convert Excel Chinese trait to English
            excel_trait_en = self.trait_zh_to_en.get(excel_trait)
            json_trait = json_monster.get("trait")

            if json_trait != excel_trait_en:
                self.error(f"{monster_id}: trait mismatch - JSON: {json_trait}, Excel: {excel_trait_en} ({excel_trait})")
                trait_mismatches += 1

        if trait_mismatches == 0:
            print(f"  ✓ All traits match\n")
        else:
            print(f"  Found {trait_mismatches} trait mismatches\n")

    def check_monster_forms(self):
        """Check if monster forms match between Excel and JSON (name-based matching)."""
        print(f"{Colors.BOLD}Checking monster forms...{Colors.RESET}")

        form_mismatches = 0

        for i, excel_data in enumerate(self.excel_monsters):
            excel_full_name = excel_data["name"]

            # Parse Excel name
            if "-" in excel_full_name:
                parts = excel_full_name.split("-", 1)
                zh_base = parts[0].strip()
                zh_form = parts[1].strip()
            else:
                zh_base = excel_full_name
                zh_form = None

            json_monster = self.json_monster_map.get((zh_base, zh_form))
            if not json_monster:
                continue

            monster_id = f"{json_monster['name']} ({excel_full_name}) [Excel row {i+2}]"

            # Skip form validation for leader monsters
            # Leader monsters in JSON may have forms to indicate evolution path,
            # but Excel doesn't specify forms for leaders
            is_leader = json_monster.get("is_leader_form", False)
            if is_leader:
                continue

            # Get JSON form
            json_form_zh = json_monster.get("localized", {}).get("zh", {}).get("form")
            json_form_en = json_monster.get("form", "default")

            # Validate form
            if zh_form is None:
                # Excel has no form (base/default form)
                if json_form_zh is not None:
                    self.error(f"{monster_id}: Excel has no form (default) but JSON has form '{json_form_zh}'")
                    form_mismatches += 1
                elif json_form_en != "default":
                    self.error(f"{monster_id}: Excel has no form but JSON form field is '{json_form_en}' (expected 'default')")
                    form_mismatches += 1
            else:
                # Excel has a form - already matched by lookup key, so they should match
                # Just verify consistency
                if json_form_zh != zh_form:
                    self.error(f"{monster_id}: Form mismatch - JSON: '{json_form_zh}', Excel: '{zh_form}'")
                    form_mismatches += 1

        if form_mismatches == 0:
            print(f"  ✓ All forms match\n")
        else:
            print(f"  Found {form_mismatches} form mismatches\n")

    def check_evolution_chains(self):
        """Check if evolution chains are consistent with Excel stage data (name-based matching)."""
        print(f"{Colors.BOLD}Checking evolution chain consistency...{Colors.RESET}")

        evolution_errors = 0

        # Build monster name lookup
        monster_names = {m["name"] for m in self.monsters}

        for i, excel_data in enumerate(self.excel_monsters):
            excel_full_name = excel_data["name"]

            # Parse Excel name
            if "-" in excel_full_name:
                parts = excel_full_name.split("-", 1)
                zh_base = parts[0].strip()
                zh_form = parts[1].strip()
            else:
                zh_base = excel_full_name
                zh_form = None

            json_monster = self.json_monster_map.get((zh_base, zh_form))
            if not json_monster:
                continue

            monster_id = f"{json_monster['name']} ({excel_full_name}) [Excel row {i+2}]"

            excel_stage = excel_data.get("stage")
            evolves_from = json_monster.get("evolves_from")

            # Parse stage (might be text like "一阶段" or number)
            stage_num = None
            if isinstance(excel_stage, int):
                stage_num = excel_stage
            elif isinstance(excel_stage, str):
                # Try to parse stage number from Chinese text
                stage_map = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5}
                for cn_num, num in stage_map.items():
                    if cn_num in excel_stage:
                        stage_num = num
                        break
                # Also try parsing as number
                if stage_num is None:
                    try:
                        stage_num = int(''.join(filter(str.isdigit, excel_stage)))
                    except (ValueError, TypeError):
                        pass

            if stage_num is None:
                continue

            # Validate evolution chain
            if stage_num == 1:
                # Base form should not have evolves_from (or it should be None)
                if evolves_from is not None:
                    self.warning(f"{monster_id}: Excel stage is 1 (base form) but JSON has evolves_from='{evolves_from}'")
            else:
                # Evolved form should have evolves_from
                if evolves_from is None:
                    self.error(f"{monster_id}: Excel stage is {stage_num} (evolved) but JSON has evolves_from=None")
                    evolution_errors += 1
                else:
                    # Check if evolves_from monster exists
                    if evolves_from not in monster_names:
                        self.error(f"{monster_id}: evolves_from '{evolves_from}' does not exist in monsters.json")
                        evolution_errors += 1

        if evolution_errors == 0:
            print(f"  ✓ All evolution chains are consistent\n")
        else:
            print(f"  Found {evolution_errors} evolution chain errors\n")

    def check_preferred_attack_style(self):
        """Check if preferred_attack_style matches base stats (all JSON monsters)."""
        print(f"{Colors.BOLD}Checking preferred_attack_style consistency...{Colors.RESET}")

        style_errors = 0

        for json_monster in self.monsters:
            zh_base = json_monster.get("localized", {}).get("zh", {}).get("name", "Unknown")
            zh_form = json_monster.get("localized", {}).get("zh", {}).get("form")
            zh_full = f"{zh_base}-{zh_form}" if zh_form else zh_base
            monster_id = f"{json_monster['name']} ({zh_full})"

            phy_atk = json_monster.get("base_phy_atk", 0)
            mag_atk = json_monster.get("base_mag_atk", 0)
            preferred = json_monster.get("preferred_attack_style")

            if phy_atk > mag_atk:
                expected = "Physical"
            elif mag_atk > phy_atk:
                expected = "Magic"
            else:
                expected = "Both"

            if preferred != expected:
                self.error(f"{monster_id}: preferred_attack_style is '{preferred}' but stats suggest '{expected}' (phy_atk={phy_atk}, mag_atk={mag_atk})")
                style_errors += 1

        if style_errors == 0:
            print(f"  ✓ All preferred_attack_style values match stats\n")
        else:
            print(f"  Found {style_errors} preferred_attack_style mismatches\n")

    def check_moveset_keys(self):
        """Check if monster moveset_key exists in monster_moves.json (all JSON monsters)."""
        print(f"{Colors.BOLD}Checking moveset_key references...{Colors.RESET}")

        missing_keys = 0
        moveset_keys = set(self.monster_moves.keys())

        for json_monster in self.monsters:
            zh_base = json_monster.get("localized", {}).get("zh", {}).get("name", "Unknown")
            zh_form = json_monster.get("localized", {}).get("zh", {}).get("form")
            zh_full = f"{zh_base}-{zh_form}" if zh_form else zh_base
            monster_id = f"{json_monster['name']} ({zh_full})"

            moveset_key = json_monster.get("moveset_key")

            if not moveset_key:
                self.error(f"{monster_id}: Missing 'moveset_key' field")
                missing_keys += 1
            elif moveset_key not in moveset_keys:
                self.error(f"{monster_id}: moveset_key '{moveset_key}' not found in monster_moves.json")
                missing_keys += 1

        if missing_keys == 0:
            print(f"  ✓ All moveset_key references are valid\n")
        else:
            print(f"  Found {missing_keys} invalid moveset_key references\n")

    def check_move_coverage(self):
        """Check if all movesets from Excel are in monster_moves.json."""
        print(f"{Colors.BOLD}Checking moveset coverage...{Colors.RESET}")

        excel_movesets = set(self.excel_moves.keys())
        json_movesets = set(self.monster_moves.keys())

        # Find movesets in Excel but not in JSON
        missing_in_json = excel_movesets - json_movesets
        if missing_in_json:
            for name in sorted(missing_in_json):
                self.warning(f"Moveset '{name}' exists in Excel but not in monster_moves.json")

        # Find movesets in JSON but not in Excel
        extra_in_json = json_movesets - excel_movesets
        if extra_in_json:
            for name in sorted(extra_in_json):
                self.info_msg(f"Moveset '{name}' exists in monster_moves.json but not in Excel")

        print(f"  Excel movesets: {len(excel_movesets)}")
        print(f"  JSON movesets: {len(json_movesets)}")
        print(f"  ✓ Coverage check complete\n")

    def check_movesets(self):
        """Check if moveset contents and order match between Excel and JSON."""
        print(f"{Colors.BOLD}Checking moveset contents and order...{Colors.RESET}")

        moveset_mismatches = 0
        skipped_empty_movesets = []

        for moveset_key, excel_moves in self.excel_moves.items():
            if moveset_key not in self.monster_moves:
                continue

            # Check if all categories are empty in Excel
            excel_learnable = excel_moves["learnable_moves"]
            excel_stones = excel_moves["move_stones"]
            excel_legacy = excel_moves["legacy_moves"]

            all_empty = (not excel_learnable and not excel_stones and not excel_legacy)
            if all_empty:
                skipped_empty_movesets.append(moveset_key)
                continue

            json_moves = self.monster_moves[moveset_key]

            # Check learnable moves (with order) - only if not empty in Excel
            if excel_learnable:
                json_learnable = json_moves.get("learnable_moves", [])

                # Position-by-position comparison
                if excel_learnable != json_learnable:
                    # First check length difference
                    if len(excel_learnable) != len(json_learnable):
                        self.error(f"Moveset '{moveset_key}': learnable moves count mismatch - Excel: {len(excel_learnable)}, JSON: {len(json_learnable)}")
                        moveset_mismatches += 1

                    # Compare position by position
                    max_len = max(len(excel_learnable), len(json_learnable))
                    position_errors = 0
                    for i in range(max_len):
                        excel_move = excel_learnable[i] if i < len(excel_learnable) else None
                        json_move = json_learnable[i] if i < len(json_learnable) else None

                        if excel_move != json_move:
                            excel_display = f"'{excel_move}'" if excel_move else "N/A"
                            json_display = f"'{json_move}'" if json_move else "N/A"
                            self.error(f"Moveset '{moveset_key}': learnable move position {i+1} - Excel: {excel_display}, JSON: {json_display}")
                            position_errors += 1
                            moveset_mismatches += 1

                    if position_errors > 0:
                        self.error(f"Moveset '{moveset_key}': {position_errors} position difference(s) in learnable moves")

            # Check move stones (with order) - only if not empty in Excel
            if excel_stones:
                json_stones = json_moves.get("move_stones", [])

                # Position-by-position comparison
                if excel_stones != json_stones:
                    # First check length difference
                    if len(excel_stones) != len(json_stones):
                        self.error(f"Moveset '{moveset_key}': move stones count mismatch - Excel: {len(excel_stones)}, JSON: {len(json_stones)}")
                        moveset_mismatches += 1

                    # Compare position by position
                    max_len = max(len(excel_stones), len(json_stones))
                    position_errors = 0
                    for i in range(max_len):
                        excel_move = excel_stones[i] if i < len(excel_stones) else None
                        json_move = json_stones[i] if i < len(json_stones) else None

                        if excel_move != json_move:
                            excel_display = f"'{excel_move}'" if excel_move else "N/A"
                            json_display = f"'{json_move}'" if json_move else "N/A"
                            self.error(f"Moveset '{moveset_key}': move stone position {i+1} - Excel: {excel_display}, JSON: {json_display}")
                            position_errors += 1
                            moveset_mismatches += 1

                    if position_errors > 0:
                        self.error(f"Moveset '{moveset_key}': {position_errors} position difference(s) in move stones")

            # Check legacy moves (with order) - only if not empty in Excel
            if excel_legacy:
                json_legacy = json_moves.get("legacy_moves", [])

                # Position-by-position comparison
                if excel_legacy != json_legacy:
                    # First check length difference
                    if len(excel_legacy) != len(json_legacy):
                        self.error(f"Moveset '{moveset_key}': legacy moves count mismatch - Excel: {len(excel_legacy)}, JSON: {len(json_legacy)}")
                        moveset_mismatches += 1

                    # Compare position by position
                    max_len = max(len(excel_legacy), len(json_legacy))
                    position_errors = 0
                    for i in range(max_len):
                        excel_move = excel_legacy[i] if i < len(excel_legacy) else None
                        json_move = json_legacy[i] if i < len(json_legacy) else None

                        if excel_move != json_move:
                            excel_display = f"'{excel_move}'" if excel_move else "N/A"
                            json_display = f"'{json_move}'" if json_move else "N/A"
                            self.error(f"Moveset '{moveset_key}': legacy move position {i+1} - Excel: {excel_display}, JSON: {json_display}")
                            position_errors += 1
                            moveset_mismatches += 1

                    if position_errors > 0:
                        self.error(f"Moveset '{moveset_key}': {position_errors} position difference(s) in legacy moves")

        # Report skipped empty movesets
        if skipped_empty_movesets:
            self.info_msg(f"Skipped {len(skipped_empty_movesets)} empty movesets (no data in Excel)")
            for moveset_key in skipped_empty_movesets:
                self.info_msg(f"  - Empty moveset: '{moveset_key}'")

        if moveset_mismatches == 0:
            print(f"  ✓ All movesets and order match\n")
        else:
            print(f"  Found moveset content/order mismatches\n")

    def print_summary(self):
        """Print summary of all checks."""
        print(f"\n{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.RESET}")
        print(f"{Colors.BOLD}{Colors.CYAN}SUMMARY{Colors.RESET}")
        print(f"{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.RESET}\n")

        if self.errors:
            print(f"{Colors.BOLD}{Colors.RED}ERRORS ({len(self.errors)}):{Colors.RESET}")
            for error in self.errors:
                print(f"  {Colors.RED}✗{Colors.RESET} {error}")
            print()

        if self.warnings:
            print(f"{Colors.BOLD}{Colors.YELLOW}WARNINGS ({len(self.warnings)}):{Colors.RESET}")
            for warning in self.warnings:
                print(f"  {Colors.YELLOW}⚠{Colors.RESET} {warning}")
            print()

        if self.info:
            print(f"{Colors.BOLD}{Colors.BLUE}INFO ({len(self.info)}):{Colors.RESET}")
            for info in self.info:
                print(f"  {Colors.BLUE}ℹ{Colors.RESET} {info}")
            print()

        # Final verdict
        print(f"{Colors.BOLD}{'='*80}{Colors.RESET}")
        if self.errors:
            print(f"{Colors.BOLD}{Colors.RED}❌ VALIDATION FAILED with {len(self.errors)} error(s){Colors.RESET}")
            exit_code = 1
        elif self.warnings:
            print(f"{Colors.BOLD}{Colors.YELLOW}⚠️  VALIDATION PASSED with {len(self.warnings)} warning(s){Colors.RESET}")
            exit_code = 0
        else:
            print(f"{Colors.BOLD}{Colors.GREEN}✅ VALIDATION PASSED - All data matches source!{Colors.RESET}")
            exit_code = 0

        return exit_code

    def save_report(self) -> Path:
        """Save validation report to a text file."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_filename = f"validation_report_{timestamp}.txt"
        report_path = self.data_dir / report_filename

        with open(report_path, 'w', encoding='utf-8') as f:
            # Header
            f.write("="*80 + "\n")
            f.write("DATA VALIDATION REPORT\n")
            f.write("="*80 + "\n\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Excel Source: {self.excel_path}\n")
            f.write(f"JSON Data Directory: {self.data_dir}\n\n")

            # Statistics
            f.write("="*80 + "\n")
            f.write("STATISTICS\n")
            f.write("="*80 + "\n\n")
            f.write(f"Total Errors:   {len(self.errors)}\n")
            f.write(f"Total Warnings: {len(self.warnings)}\n")
            f.write(f"Total Info:     {len(self.info)}\n\n")
            f.write(f"Excel Monsters: {len(self.excel_monsters)}\n")
            f.write(f"JSON Monsters:  {len(self.monsters)}\n\n")

            # Errors
            if self.errors:
                f.write("="*80 + "\n")
                f.write(f"ERRORS ({len(self.errors)})\n")
                f.write("="*80 + "\n\n")
                for i, error in enumerate(self.errors, 1):
                    f.write(f"{i}. {error}\n")
                f.write("\n")

            # Warnings
            if self.warnings:
                f.write("="*80 + "\n")
                f.write(f"WARNINGS ({len(self.warnings)})\n")
                f.write("="*80 + "\n\n")
                for i, warning in enumerate(self.warnings, 1):
                    f.write(f"{i}. {warning}\n")
                f.write("\n")

            # Info
            if self.info:
                f.write("="*80 + "\n")
                f.write(f"INFO ({len(self.info)})\n")
                f.write("="*80 + "\n\n")
                for i, info in enumerate(self.info, 1):
                    f.write(f"{i}. {info}\n")
                f.write("\n")

            # Final verdict
            f.write("="*80 + "\n")
            f.write("FINAL VERDICT\n")
            f.write("="*80 + "\n\n")
            if self.errors:
                f.write(f"❌ VALIDATION FAILED with {len(self.errors)} error(s)\n")
            elif self.warnings:
                f.write(f"⚠️  VALIDATION PASSED with {len(self.warnings)} warning(s)\n")
            else:
                f.write("✅ VALIDATION PASSED - All data matches source!\n")

        return report_path


def main():
    """Main entry point."""
    script_dir = Path(__file__).parent
    data_dir = script_dir.parent.parent / "data"
    excel_path = data_dir / "data_all.xlsx"

    if not data_dir.exists():
        print(f"{Colors.RED}Error: Data directory not found at {data_dir}{Colors.RESET}")
        return 1

    if not excel_path.exists():
        print(f"{Colors.RED}Error: Excel file not found at {excel_path}{Colors.RESET}")
        return 1

    print(f"{Colors.BOLD}Data directory: {data_dir}{Colors.RESET}")
    print(f"{Colors.BOLD}Excel file: {excel_path}{Colors.RESET}\n")

    checker = SourceCorrectnessChecker(data_dir, excel_path)
    exit_code = checker.check_all()

    return exit_code


if __name__ == "__main__":
    sys.exit(main())
